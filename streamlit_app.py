"""Streamlit UI for flow-bof-automation (Phase 2).

Wraps the existing CLI. Light operations (reading the CSV, parsing the
manifest, listing reference images, reading log files) call the
existing Python functions directly. Heavy operations that touch the
browser or rewrite the CSV subprocess `python main.py …` so the same
code path the CLI uses also runs from the UI — no duplicated business
logic.

The UI is single-threaded; while a Generate Images or Sync Favorites
subprocess is running, the page can't be navigated until it finishes.
Output streams live into the page via a placeholder.
"""

from __future__ import annotations

import json
import os
import subprocess
import sys
import urllib.error
import urllib.request
from collections import Counter
from datetime import datetime
from pathlib import Path

import streamlit as st


# Make the project root importable even when run from elsewhere.
ROOT = Path(__file__).resolve().parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from src.config import load_settings
from src.csv_workflow import (
    STATUS_DONE,
    STATUS_IMAGE_APPROVED,
    STATUS_IMAGE_REJECTED,
    STATUS_IMAGE_SUBMITTED,
    STATUS_PENDING,
    STATUS_VIDEO_PENDING,
    STATUS_VIDEO_SUBMITTED,
    load_csv,
    media_ids_of,
)
from src.manifest_workflow import parse_manifest, resolve_reference_image
from src.batch_workflow import (
    PRODUCT_STATUS_DRAFT,
    PRODUCT_STATUS_EXPORTED,
    PRODUCT_STATUS_READY,
    REFERENCE_ROLES,
    ProductCard,
    add_to_image_pool,
    attach_pool_image_to_product,
    clean_product_title,
    create_batch,
    detect_url_source,
    export_manifest,
    image_pool_dir,
    list_batches,
    list_image_pool,
    load_products as load_batch_products,
    new_product_id,
    reference_image_rel,
    remove_from_image_pool,
    save_products as save_batch_products,
    save_reference_image,
)
from ai.prompt_generator import (
    KNOWN_PROVIDERS,
    extract_json,
    get_provider,
    test_ai_provider,
    validate_ai_output,
)
from src.unmatched_favorites import (
    UnmatchedFavorite,
    load_unmatched as load_unmatched_favorites,
    remove_unmatched as remove_unmatched_favorite,
)
from src.sync_workflow import bind_unmatched_favorite_to_product
from src.user_settings import (
    SETTINGS_FILE as USER_SETTINGS_FILE,
    SECRETS_FILE as USER_SECRETS_FILE,
    UserSecrets,
    UserSettings as UISettings,
    apply_to_env as apply_user_settings_to_env,
    load_secrets as load_user_secrets,
    load_settings as load_user_settings,
    mask_key,
    save_secrets as save_user_secrets,
    save_settings as save_user_settings,
)
from src import health as health_checks


# Curated UK retailer list — exactly the retailers covered in the
# UK_SYSTEM_PROMPT mapping table. The product card surfaces this as a
# dropdown when MARKET=UK so the user can override the AI's pick (or
# pre-seed before running AI prompts).
UK_RETAILERS = [
    "Boots",
    "Sephora UK",
    "Selfridges",
    "Holland & Barrett",
    "Primark",
    "Schuh",
    "JD Sports",
    "IKEA",
    "John Lewis",
    "Currys",
    "Argos",
    "Smyths Toys",
    "Pets at Home",
    "Tesco",
]


def _active_market() -> str:
    m = (os.environ.get("MARKET") or "US").strip().upper()
    return m if m in ("US", "UK") else "US"


def _prompt_mode_label() -> str:
    """User-facing label for the active image-prompt style."""
    if _active_market() == "UK":
        return "UK Retail Store Display"
    return "US Retail Editorial (AIBOF)"

# Apply UI-saved settings/secrets to this process's os.environ before
# any provider class reads from it.
apply_user_settings_to_env()

# Optional clipboard paste button. Degrades gracefully if the package
# isn't installed or the browser refuses clipboard access.
try:
    from streamlit_paste_button import paste_image_button as _paste_image_button
    HAS_PASTE_BUTTON = True
except Exception:  # noqa: BLE001
    _paste_image_button = None
    HAS_PASTE_BUTTON = False


SETTINGS = load_settings()
REPO_ROOT = SETTINGS.repo_root
PRODUCTS_CSV = SETTINGS.products_csv
MANIFEST_PATH = SETTINGS.manifest_path
REFERENCE_IMAGES_DIR = SETTINGS.reference_images_dir
LOGS_DIR = SETTINGS.logs_dir
IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".webp"}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


import re
import time


# Friendly label for each CLI flag the UI invokes. Shown at the top of
# the running-status placeholder instead of the raw argv.
_CLI_FRIENDLY_LABELS = {
    "--check-browser":                        "Checking browser connection",
    "--validate-manifest":                    "Validating manifest",
    "--load-manifest":                        "Loading manifest",
    "--generate-images":                      "Generating product images",
    "--sync-favorites":                       "Scanning favorited images",
    "--generate-videos":                      "Generating videos from favorited images",
    "--generate-videos-from-favorited-tiles": "Generating videos from favorited images",
    "--generate-videos-from-approved-rows":   "Generating videos from approved rows (legacy)",
    "--list-unmatched-favorites":             "Scanning favorited images",
    "--capture-tiles":                        "Capturing tile IDs",
    "--scan-images":                          "Scanning reference images",
    "--list-status":                          "Listing run status",
    "--bind-favorite":                        "Binding favorite to product",
    "--debug-selectors":                      "Debugging selectors",
}


def _friendly_label_for(args: list[str]) -> str:
    for a in args:
        if a in _CLI_FRIENDLY_LABELS:
            return _CLI_FRIENDLY_LABELS[a]
    return "Running command"


# Per-line patterns the runner uses to lift friendly progress events out
# of the otherwise-noisy structured log. We deliberately ignore the
# Playwright-tile-level chatter (hovered/clicked/menu visible) — those
# are debug noise; the user only wants per-item progress.
_PROGRESS_PATTERNS = [
    # Video: favorited-tile mode.
    (re.compile(r"Scanned (\d+) tile\(s\); (\d+) favorited image tile\(s\) eligible\."),
     lambda m: f"Found {m.group(2)} favorited image{'s' if int(m.group(2)) != 1 else ''}."),
    (re.compile(r"Animating (\d+) favorited tile\(s\)"),
     lambda m: f"Submitting {m.group(1)} video{'s' if int(m.group(1)) != 1 else ''}..."),
    (re.compile(r"--- \[(\d+)/(\d+)\] media_id=\S+"),
     lambda m: f"Submitting video {m.group(1)} of {m.group(2)}..."),
    (re.compile(r"Video submitted for media_id=\S+"),
     lambda m: "Video submitted."),
    (re.compile(r"Skipping media_id=\S+ \(already submitted"),
     lambda m: "Skipped (already submitted)."),
    (re.compile(r"Video failed for media_id=\S+"),
     lambda m: "One video failed; continuing batch."),
    # Image generation.
    (re.compile(r"Processing (\d+) row\(s\)"),
     lambda m: f"Generating {m.group(1)} image{'s' if int(m.group(1)) != 1 else ''}..."),
    (re.compile(r"--- \[(\d+)/(\d+)\] id=\S+"),
     lambda m: f"Submitting image {m.group(1)} of {m.group(2)}..."),
    # Sync favorites.
    (re.compile(r"Scanning Flow Labs gallery for favorited tiles"),
     lambda m: "Scanning Flow Labs gallery..."),
    (re.compile(r"Hearted (\d+) tile"),
     lambda m: f"Hearted: {m.group(1)} tile(s)."),
]


def _parse_progress(line: str) -> str | None:
    """If a line matches a known event, return a one-line friendly form."""
    for pat, fmt in _PROGRESS_PATTERNS:
        m = pat.search(line)
        if m:
            return fmt(m)
    return None


# Final-summary patterns: after the run completes, we re-scan the full
# output for these and lift counts into the summary card.
def _parse_command_summary(args: list[str], output: str, elapsed_s: float) -> dict[str, str]:
    """Extract a clean summary dict for the UI summary card.

    Best-effort: any value we can't pull from the logs is simply
    omitted. Always returns at least 'Elapsed'.
    """
    summary: dict[str, str] = {"Elapsed": f"{elapsed_s:.1f}s"}

    # Mode (safe / balanced / fast).
    m = re.search(r"\(mode=(\w+),", output)
    if m:
        summary["Mode"] = m.group(1)

    # --- Video: favorited-tile mode ----------------------------------
    m = re.search(
        r"Favorited-tile video batch done in [\d.]+s — "
        r"(\d+) submitted, (\d+) failed, (\d+) skipped",
        output,
    )
    if m:
        summary["Videos submitted"] = m.group(1)
        summary["Failed"] = m.group(2)
        summary["Already submitted (skipped)"] = m.group(3)
        if "Mode" in summary:
            summary["Prompt"] = "Universal blanket prompt"
    else:
        # No final-summary line — fall back to counting per-line events.
        # Useful when the run was interrupted.
        scanned = re.search(r"Scanned \d+ tile\(s\); (\d+) favorited", output)
        if scanned:
            summary.setdefault("Favorited images found", scanned.group(1))

    # --- Video: approved-rows (legacy) -------------------------------
    m = re.search(
        r"Video batch done in [\d.]+s — (\d+) submitted, (\d+) failed",
        output,
    )
    if m and "Videos submitted" not in summary:
        summary["Videos submitted"] = m.group(1)
        summary["Failed"] = m.group(2)

    # --- Image generation --------------------------------------------
    m = re.search(
        r"Image batch done in [\d.]+s — (\d+) submitted, (\d+) failed",
        output,
    )
    if m:
        summary["Images submitted"] = m.group(1)
        summary["Failed"] = m.group(2)
    # Also catch the per-row "skipped" counter if it appears.
    skipped_imgs = len(re.findall(r"Skipping row id=\S+ \(", output))
    if skipped_imgs:
        summary["Skipped"] = str(skipped_imgs)

    # --- Sync favorites ----------------------------------------------
    m = re.search(r"Hearted (\d+) tile", output)
    if m:
        summary["Hearted in Flow"] = m.group(1)
    m = re.search(r"Approved (\d+) row\(s\)", output)
    if m:
        summary["Approved"] = m.group(1)
    m = re.search(r"(\d+) unmatched favorite", output)
    if m:
        summary["Need review"] = m.group(1)

    return summary


def _last_meaningful_lines(output: str, n: int = 10) -> list[str]:
    """Pick the last N non-trivial log lines for the recap section."""
    lines = [ln.rstrip() for ln in output.splitlines() if ln.strip()]
    return lines[-n:]


# ===========================================================================
# Phase-2b: optional HTTP execution mode against the local agent server.
#
# Set AGENT_EXECUTION_MODE=http in Setup to route SUPPORTED commands at the
# /jobs/run-stream endpoint instead of subprocessing main.py. The two
# transports return the same shape ((exit_code, output)), so the rest of
# the UI is transport-agnostic.
#
# Commands that aren't a 1:1 match for an agent job (manifest mutations,
# legacy CSV sync) always stay on the CLI path even in HTTP mode.
# ===========================================================================


def _agent_execution_mode() -> str:
    """'http' or 'cli'. Unknown values fall through to 'cli'."""
    raw = (os.environ.get("AGENT_EXECUTION_MODE") or "cli").strip().lower()
    return "http" if raw == "http" else "cli"


def _agent_base_url() -> str:
    raw = (os.environ.get("AGENT_BASE_URL") or "http://127.0.0.1:9444").strip()
    return raw.rstrip("/") or "http://127.0.0.1:9444"


def _agent_http_headers() -> dict:
    token = (os.environ.get("AGENT_API_TOKEN") or "").strip()
    return {"Authorization": f"Bearer {token}"} if token else {}


def agent_http_health(timeout: float = 5.0) -> tuple[bool, dict]:
    """Cheap reachability probe — calls GET /health on the agent.

    Returns (reachable, payload). On any transport error, reachable is
    False and payload carries the error so the UI can render it.
    """
    import httpx
    try:
        r = httpx.get(
            f"{_agent_base_url()}/health",
            headers=_agent_http_headers(),
            timeout=timeout,
        )
        if r.status_code == 200:
            return True, r.json()
        return False, {
            "http_status": r.status_code,
            "body":        (r.text or "")[:400],
        }
    except Exception as exc:  # noqa: BLE001
        return False, {
            "error":   f"{type(exc).__name__}: {exc}",
            "url":     f"{_agent_base_url()}/health",
        }


# CLI args → agent job envelope. Only the subset we've migrated:
# --check-browser, --generate-videos, --list-unmatched-favorites.
# Other args return None so run_cli stays on the subprocess path.
def _args_to_agent_job(args: list[str]) -> dict | None:
    if not args:
        return None
    first = args[0]

    if first == "--check-browser":
        return {
            "job_type": "health_check",
            "payload":  {},
        }

    if first in ("--list-unmatched-favorites",):
        # Read-only Flow scan — same as the agent's scan_favorited_images,
        # which is what the user actually wants when clicking Scan
        # Favorites from the BOF page.
        return {
            "job_type": "scan_favorited_images",
            "payload":  {"limit": 100,
                         "include_non_favorites": False,
                         "include_videos": False},
        }

    if first == "--generate-videos":
        # --generate-videos --limit N [--include-already-submitted]
        limit = 30
        include_already = False
        i = 1
        while i < len(args):
            if args[i] == "--limit" and i + 1 < len(args):
                try:
                    limit = int(args[i + 1])
                except ValueError:
                    pass
                i += 2
            elif args[i] == "--include-already-submitted":
                include_already = True
                i += 1
            else:
                # Unrecognized extra arg — stay safe, drop to CLI.
                return None
        return {
            "job_type": "generate_flow_videos_from_favorites",
            "payload":  {
                "limit": limit,
                "include_already_submitted": include_already,
            },
        }

    return None


def _summary_from_envelope(envelope: dict) -> dict[str, str]:
    """Lift the friendly summary fields from a successful agent envelope.

    The envelope already has structured counts in result.<...>; no log
    regexing needed.
    """
    summary: dict[str, str] = {}
    result = envelope.get("result") or {}
    elapsed = result.get("elapsed_seconds")
    if elapsed is not None:
        summary["Elapsed"] = f"{float(elapsed):.1f}s"

    # Whitelist of result keys that make sense as user-facing chips.
    # Anything else stays in the "Show technical details" expander.
    pretty = {
        "favorited_images_found":     "Favorited found",
        "submitted":                  "Submitted",
        "failed":                     "Failed",
        "skipped_already_submitted":  "Skipped (already submitted)",
        "favorited_image_count":      "Favorited images",
        "tile_count":                 "Tiles scanned",
        "items_received":             "Items received",
        "processed":                  "Processed",
        "flow_reachable":             "Flow reachable",
        "chrome_reachable":           "Chrome reachable",
    }
    for raw_key, label in pretty.items():
        if raw_key in result:
            summary[label] = str(result[raw_key])
    return summary


def _run_via_agent_http(args: list[str], job: dict) -> tuple[int, str]:
    """Send the job to the local agent's /jobs/run-stream endpoint and
    render progress events in the same UI cadence as the subprocess
    path. Public contract is unchanged: returns (exit_code, output)."""
    import httpx
    import uuid

    label = _friendly_label_for(args) + " (via local agent)"
    status_box = st.empty()
    progress_box = st.empty()
    status_box.info(f"**{label}** — starting…")

    job_envelope = {
        "protocol_version": "0.1",
        "job_id":           f"streamlit-{uuid.uuid4().hex[:12]}",
        **job,
    }
    url = f"{_agent_base_url()}/jobs/run-stream"
    headers = {**_agent_http_headers(), "Content-Type": "application/json"}

    cli_started = time.monotonic()
    final_envelope: dict | None = None
    progress_lines: list[str] = []
    raw_event_lines: list[str] = []  # preserved for the technical-details expander
    last_friendly = ""

    try:
        with st.spinner(f"{label}…"):
            with httpx.stream(
                "POST", url,
                json=job_envelope,
                headers=headers,
                timeout=None,  # streaming endpoint, can be many minutes
            ) as response:
                if response.status_code != 200:
                    body = response.read().decode("utf-8", errors="replace")[:400]
                    raise httpx.HTTPStatusError(
                        f"HTTP {response.status_code}: {body}",
                        request=response.request,
                        response=response,
                    )
                for line in response.iter_lines():
                    if not line.strip():
                        continue
                    raw_event_lines.append(line)
                    try:
                        evt = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    et = evt.get("event_type")
                    if et == "result":
                        final_envelope = evt.get("envelope")
                    elif et == "progress":
                        msg = evt.get("message") or evt.get("stage") or ""
                        cur = evt.get("current")
                        tot = evt.get("total")
                        if cur is not None and tot is not None:
                            friendly = f"• {msg} ({cur}/{tot})"
                        else:
                            friendly = f"• {msg}"
                        if friendly != last_friendly:
                            last_friendly = friendly
                            progress_lines.append(friendly)
                            # Show the last ~3 lines so the user
                            # sees current activity, not a wall of text.
                            progress_box.write("\n".join(progress_lines[-3:]))
    except httpx.HTTPStatusError as exc:
        elapsed = time.monotonic() - cli_started
        progress_box.empty()
        status_box.error(
            f"**{label}** — HTTP error after {elapsed:.1f}s: {exc}"
        )
        return 1, str(exc)
    except httpx.HTTPError as exc:
        elapsed = time.monotonic() - cli_started
        progress_box.empty()
        status_box.error(
            f"**{label}** — local agent unreachable at "
            f"`{_agent_base_url()}` after {elapsed:.1f}s. "
            f"Start it with: `docker compose --profile agent up -d agent`. "
            f"({type(exc).__name__}: {exc})"
        )
        return 1, str(exc)

    elapsed = time.monotonic() - cli_started
    progress_box.empty()

    if final_envelope is None:
        status_box.error(
            f"**{label}** — stream ended without a result envelope."
        )
        return 1, "(no envelope received)"

    success = final_envelope.get("status") == "succeeded"
    if success:
        status_box.success(f"**{label}** — done in {elapsed:.1f}s.")
    else:
        err = final_envelope.get("error") or {}
        status_box.error(
            f"**{label}** — job failed "
            f"(code={err.get('code', '?')}): "
            f"{(err.get('message') or '')[:200]}"
        )

    summary = _summary_from_envelope(final_envelope)
    if summary:
        progress_box.markdown(
            "  ·  ".join(f"**{k}:** {v}" for k, v in summary.items())
        )

    with st.expander("Show technical details", expanded=False):
        st.caption(f"`POST {url}`")
        st.code(json.dumps(final_envelope, indent=2), language="json")
        if raw_event_lines:
            st.caption(f"NDJSON stream ({len(raw_event_lines)} line(s))")
            st.code("\n".join(raw_event_lines), language="json")

    output = json.dumps(final_envelope, indent=2)
    st.session_state["last_command_output"]   = output
    st.session_state["last_command_args"]     = args
    st.session_state["last_command_elapsed_s"] = elapsed
    st.session_state["last_command_exit_code"] = 0 if success else 1
    st.session_state["last_command_summary"]  = summary
    st.session_state["last_command_label"]    = label
    st.session_state["last_command_cmd"]      = (
        f"POST {url} (job_type={final_envelope.get('job_type')})"
    )
    return (0 if success else 1, output)


def run_cli(args: list[str]) -> tuple[int, str]:
    """Run an agent action against the local agent HTTP API (if
    AGENT_EXECUTION_MODE=http AND the action maps to an agent job), or
    fall back to subprocessing ``python main.py <args>``.

    Public contract is unchanged: returns ``(exit_code, full_output)``.
    """
    # Phase-2b: try the local agent HTTP path first. If the user hasn't
    # opted in (or the args aren't agent-mappable), fall straight through
    # to the legacy subprocess implementation below.
    if _agent_execution_mode() == "http":
        job = _args_to_agent_job(args)
        if job is not None:
            return _run_via_agent_http(args, job)

    cmd = [sys.executable, str(ROOT / "main.py"), *args]
    label = _friendly_label_for(args)

    # Per-section state lives in the page DOM via st.empty placeholders.
    # We wrap the subprocess in a st.spinner so Streamlit shows its
    # running indicator immediately after click — bridges the visual
    # gap between "page re-runs from top" and "status_box first paints"
    # which is what users perceive as "the page turned white".
    status_box = st.empty()
    progress_box = st.empty()
    status_box.info(f"**{label}** — starting…")

    cli_started = time.monotonic()
    try:
        proc = subprocess.Popen(
            cmd,
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            bufsize=1,
            cwd=str(ROOT),
            env=os.environ.copy(),
        )
    except FileNotFoundError as exc:
        status_box.error(f"Could not start subprocess: {exc}")
        return 1, str(exc)

    assert proc.stdout is not None
    raw_lines: list[str] = []
    last_friendly = ""
    with st.spinner(f"{label}…"):
        for line in iter(proc.stdout.readline, ""):
            raw_lines.append(line)
            friendly = _parse_progress(line)
            if friendly and friendly != last_friendly:
                last_friendly = friendly
                progress_box.write(f"• {friendly}")
        proc.wait()
    elapsed = time.monotonic() - cli_started
    output = "".join(raw_lines)

    summary = _parse_command_summary(args, output, elapsed)

    if proc.returncode == 0:
        status_box.success(f"**{label}** — done in {elapsed:.1f}s.")
    else:
        status_box.error(
            f"**{label}** — failed (exit {proc.returncode}, "
            f"elapsed {elapsed:.1f}s). See technical logs below."
        )

    # Render the summary as a single dense markdown line instead of
    # st.metric tiles. The metric layout stretches its columns to fill
    # the row, so with only 2-3 summary keys you get a wide expanse of
    # whitespace either side. A pill-style caption packs the same info
    # into one row with no padding.
    if summary:
        progress_box.markdown(
            "  ·  ".join(f"**{k}:** {v}" for k, v in summary.items())
        )
    else:
        # No parsed summary → keep the per-line progress events visible
        # so the run isn't represented by a single success banner alone.
        pass

    # Recap: last few meaningful log lines, unfiltered. Only useful when
    # our parser couldn't extract a structured summary — otherwise it's
    # redundant with the technical-logs expander and adds vertical
    # whitespace for no information gain.
    if not summary:
        tail = _last_meaningful_lines(output, 8)
        if tail:
            with st.expander("Recent activity", expanded=False):
                st.code("\n".join(tail), language="text")

    with st.expander("Show technical logs", expanded=False):
        st.caption(f"`{' '.join(cmd)}`")
        st.code(output or "(no output)", language="text")

    st.session_state["last_command_output"] = output
    st.session_state["last_command_args"] = args
    st.session_state["last_command_elapsed_s"] = elapsed
    st.session_state["last_command_exit_code"] = proc.returncode
    st.session_state["last_command_summary"] = summary
    st.session_state["last_command_label"] = label
    st.session_state["last_command_cmd"] = " ".join(cmd)
    return proc.returncode, output


def chrome_cdp_url() -> str:
    return os.environ.get("CHROME_CDP_URL", "http://cdp-proxy:9333")


def chrome_reachable() -> tuple[bool, str]:
    url = f"{chrome_cdp_url()}/json/version"
    try:
        with urllib.request.urlopen(url, timeout=3) as resp:
            return True, resp.read().decode("utf-8", errors="replace")
    except urllib.error.URLError as exc:
        return False, str(exc)
    except Exception as exc:  # noqa: BLE001
        return False, repr(exc)


def csv_status_counts() -> tuple[int, dict[str, int]]:
    if not PRODUCTS_CSV.exists():
        return 0, {}
    try:
        rows = load_csv(PRODUCTS_CSV)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Failed to read {PRODUCTS_CSV}: {exc}")
        return 0, {}
    return len(rows), dict(Counter(r.status for r in rows))


def list_reference_images() -> list[Path]:
    if not REFERENCE_IMAGES_DIR.exists():
        return []
    return sorted(
        p for p in REFERENCE_IMAGES_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTS
    )


# ---------------------------------------------------------------------------
# Pages
# ---------------------------------------------------------------------------


def _ai_provider_is_configured() -> bool:
    """True when the active AI provider has credentials. Manual counts."""
    name = (os.environ.get("AI_PROVIDER") or "").strip().lower()
    if not name:
        return False
    if name == "manual":
        return True
    try:
        provider = get_provider(name)
        ok, _msg = provider.is_configured()
        return ok
    except Exception:  # noqa: BLE001
        return False


def _status_icon(status: str) -> str:
    return {"ok": "✅", "warn": "⚠️", "fail": "❌"}.get(status, "•")


def _render_health_checklist(items: list[dict]) -> None:
    for item in items:
        icon = _status_icon(item["status"])
        st.markdown(
            f"{icon} **{item['label']}** — {item['message']}"
        )


def _provider_default_model(name: str) -> str:
    return {
        "openai":     "gpt-4o-mini",
        "anthropic":  "claude-3-5-sonnet-latest",
        "openrouter": "",  # empty = openrouter/auto
    }.get(name, "")


def _render_execution_mode_section(saved_settings, saved_secrets) -> None:
    """Setup-page block for Phase-2b execution mode.

    Lets the user pick whether agent-supported actions go through the
    local agent's HTTP API or through the legacy CLI subprocess. The
    CLI mode remains the default — switching is purely opt-in.

    Surfaces a live status indicator so the user knows whether the
    agent is reachable before they switch into HTTP mode.
    """
    st.subheader("Execution mode (advanced)")
    st.caption(
        "Most actions in this UI subprocess `python main.py …`. If you "
        "boot the local agent HTTP API (`docker compose --profile agent "
        "up -d agent`), you can switch agent-supported actions to talk "
        "to it directly for cleaner progress and a stepping stone to a "
        "future SaaS dashboard. CLI mode is the default and always "
        "works."
    )

    current_mode = (
        saved_settings.agent_execution_mode
        or os.environ.get("AGENT_EXECUTION_MODE")
        or "cli"
    ).strip().lower()
    if current_mode not in ("cli", "http"):
        current_mode = "cli"
    mode_options = ["cli", "http"]
    mode_idx = mode_options.index(current_mode)

    # Live agent reachability probe.
    if st.button("Test Agent API", key="setup_test_agent_api"):
        with st.spinner("Calling /health on the local agent…"):
            ok, info = agent_http_health(timeout=4.0)
        if ok:
            r = info.get("result") or {}
            st.success(
                f"Local Agent API is reachable at `{_agent_base_url()}`. "
                f"agent_ok={r.get('agent_ok')}, "
                f"chrome_reachable={r.get('chrome_reachable')}, "
                f"flow_reachable={r.get('flow_reachable')}."
            )
        else:
            st.error(
                f"Local Agent API is **not reachable** at "
                f"`{_agent_base_url()}`. "
                "Start it with: `docker compose --profile agent up -d agent`"
            )
            with st.expander("Show technical details", expanded=False):
                st.code(json.dumps(info, indent=2), language="json")

    # Status chip — every render, cached for ~30s in session_state so we
    # don't slam /health on every rerun.
    status_cache_key = "agent_http_status_cache"
    cache = st.session_state.get(status_cache_key) or {}
    cache_age = time.monotonic() - cache.get("at", 0.0)
    if not cache or cache_age > 30.0:
        ok, info = agent_http_health(timeout=2.0)
        cache = {"ok": ok, "info": info, "at": time.monotonic()}
        st.session_state[status_cache_key] = cache
    if cache["ok"]:
        st.success(
            f"Agent API: **Connected** ({_agent_base_url()})"
        )
    else:
        st.warning(
            f"Agent API: **Not connected** at {_agent_base_url()}. "
            "Switch off HTTP mode below if you don't plan to start it, "
            "or run `docker compose --profile agent up -d agent`."
        )

    with st.form("execution_mode_form"):
        new_mode = st.selectbox(
            "Execution mode",
            mode_options,
            index=mode_idx,
            format_func=lambda m: {"cli": "Local CLI (default)",
                                   "http": "Local Agent HTTP"}[m],
            help=(
                "Local CLI subprocesses `python main.py …` exactly as "
                "today. Local Agent HTTP routes agent-supported actions "
                "to http://127.0.0.1:9444 — actions not yet migrated "
                "(manifest fresh-load, etc.) still fall through to CLI."
            ),
        )
        agent_url = st.text_input(
            "Agent base URL",
            value=saved_settings.agent_base_url or "http://127.0.0.1:9444",
            help=(
                "Default http://127.0.0.1:9444. Change this only if you "
                "moved the agent server to a different host/port."
            ),
        )
        agent_token = st.text_input(
            "Agent API token (optional)",
            value=saved_secrets.agent_api_token,
            type="password",
            help=(
                f"Currently saved: {mask_key(saved_secrets.agent_api_token)}. "
                "Required only if the agent itself was started with "
                "AGENT_API_TOKEN set."
            ),
        )
        save_exec = st.form_submit_button("💾  Save execution mode", type="primary")

    if save_exec:
        try:
            new_settings = load_user_settings()
            new_settings.agent_execution_mode = new_mode.strip().lower()
            new_settings.agent_base_url = agent_url.strip()
            new_secrets = load_user_secrets()
            new_secrets.agent_api_token = agent_token.strip()
            save_user_settings(new_settings)
            save_user_secrets(new_secrets)
            apply_user_settings_to_env(new_settings, new_secrets)
            # Force the status cache to refresh on the next render.
            st.session_state.pop(status_cache_key, None)
            st.success(
                f"Saved. Execution mode = **{new_mode}**. "
                "Take effect immediately for new actions."
            )
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not save execution mode: {exc}")


def page_setup() -> None:
    st.title("Setup")
    st.caption(
        "One place to verify the environment and configure your AI provider. "
        "Settings save to ``data/settings.local.json`` and ``data/secrets.local.json`` "
        "— never committed to git, never baked into the Docker image."
    )

    # --- Health checklist -------------------------------------------------
    st.subheader("Health check")
    st.caption("Re-run to refresh after starting Chrome or saving an API key.")
    if st.button("🩺  Run health check", key="run_health_check"):
        st.session_state["health_results"] = health_checks.run_all_checks()
    results = st.session_state.get("health_results")
    if results is None:
        results = health_checks.run_all_checks()
        st.session_state["health_results"] = results
    _render_health_checklist(results)

    st.divider()

    # --- AI Settings ------------------------------------------------------
    st.subheader("AI provider settings")

    saved_settings = load_user_settings()
    saved_secrets = load_user_secrets()

    current_provider = (
        saved_settings.ai_provider
        or os.environ.get("AI_PROVIDER")
        or "manual"
    ).strip().lower()
    try:
        prov_idx = list(KNOWN_PROVIDERS).index(current_provider)
    except ValueError:
        prov_idx = list(KNOWN_PROVIDERS).index("manual")

    # Resolve the saved market into a dropdown index.
    saved_market = (
        saved_settings.market or os.environ.get("MARKET") or "US"
    ).strip().upper()
    if saved_market not in ("US", "UK"):
        saved_market = "US"
    market_idx = ["US", "UK"].index(saved_market)

    with st.form("ai_settings_form"):
        provider_choice = st.selectbox(
            "AI Provider",
            list(KNOWN_PROVIDERS),
            index=prov_idx,
            help=(
                "openai → uses OPENAI_API_KEY.\n"
                "anthropic → uses ANTHROPIC_API_KEY.\n"
                "openrouter → uses OPENROUTER_API_KEY.\n"
                "manual → no API call; you write the prompts."
            ),
        )
        market_choice = st.selectbox(
            "Market",
            ["US", "UK"],
            index=market_idx,
            help=(
                "US → original AIBOF editorial framework (DISPLAY METHOD, "
                "LIGHTING SENTENCE, big-box US retailers, no real brand "
                "names).\n"
                "UK → Apex Initiative UK retail prompt library — one short "
                "sentence dropping the product into the right UK store "
                "(Boots, Sephora UK, Selfridges, Primark, Currys, etc.).\n"
                "Video prompts are the same in both markets (the universal "
                "blanket prompt)."
            ),
        )

        col_a, col_b = st.columns(2)
        with col_a:
            st.markdown("**OpenAI**")
            openai_key = st.text_input(
                "OpenAI API key",
                value=saved_secrets.openai_api_key,
                type="password",
                help=f"Currently saved: {mask_key(saved_secrets.openai_api_key)}",
            )
            openai_model = st.text_input(
                "OpenAI model",
                value=saved_settings.openai_model or _provider_default_model("openai"),
            )

            st.markdown("**Anthropic**")
            anthropic_key = st.text_input(
                "Anthropic API key",
                value=saved_secrets.anthropic_api_key,
                type="password",
                help=f"Currently saved: {mask_key(saved_secrets.anthropic_api_key)}",
            )
            anthropic_model = st.text_input(
                "Anthropic model",
                value=saved_settings.anthropic_model or _provider_default_model("anthropic"),
            )

        with col_b:
            st.markdown("**OpenRouter**")
            openrouter_key = st.text_input(
                "OpenRouter API key",
                value=saved_secrets.openrouter_api_key,
                type="password",
                help=f"Currently saved: {mask_key(saved_secrets.openrouter_api_key)}",
            )
            openrouter_model = st.text_input(
                "OpenRouter model",
                value=saved_settings.openrouter_model,
                help=(
                    "Leave blank to use OpenRouter's auto-router "
                    "(openrouter/auto). Lock to e.g. "
                    "anthropic/claude-3.5-sonnet for predictable cost."
                ),
            )
            openrouter_site_url = st.text_input(
                "OpenRouter site URL (optional)",
                value=saved_settings.openrouter_site_url,
            )
            openrouter_app_name = st.text_input(
                "OpenRouter app name (optional)",
                value=saved_settings.openrouter_app_name,
            )

        col_save, col_test = st.columns([1, 1])
        with col_save:
            submitted_save = st.form_submit_button("💾  Save settings", type="primary")
        with col_test:
            submitted_test = st.form_submit_button("🧪  Test API key")

    if submitted_save or submitted_test:
        # Validate non-manual provider has its own key.
        provider_key_lookup = {
            "openai":     openai_key,
            "anthropic":  anthropic_key,
            "openrouter": openrouter_key,
        }
        if provider_choice != "manual" and not provider_key_lookup.get(provider_choice, "").strip():
            st.warning(
                f"You picked **{provider_choice}** but didn't enter a "
                f"{provider_choice.upper()}_API_KEY. The AI step will "
                f"fail until you do."
            )

        new_settings = UISettings(
            ai_provider=provider_choice,
            openai_model=openai_model.strip(),
            anthropic_model=anthropic_model.strip(),
            openrouter_model=openrouter_model.strip(),
            openrouter_site_url=openrouter_site_url.strip(),
            openrouter_app_name=openrouter_app_name.strip(),
            market=market_choice,
            # Preserve previously saved video knobs so saving the AI
            # form doesn't clobber blanket-prompt and source-mode picks.
            use_blanket_video_prompt=saved_settings.use_blanket_video_prompt,
            blanket_video_prompt=saved_settings.blanket_video_prompt,
            video_source_mode=saved_settings.video_source_mode,
        )
        new_secrets = UserSecrets(
            openai_api_key=openai_key.strip(),
            anthropic_api_key=anthropic_key.strip(),
            openrouter_api_key=openrouter_key.strip(),
        )

        if submitted_save:
            try:
                save_user_settings(new_settings)
                save_user_secrets(new_secrets)
                apply_user_settings_to_env(new_settings, new_secrets)
                st.session_state["health_results"] = health_checks.run_all_checks()
                st.success(
                    f"Saved. Provider: **{provider_choice}**. "
                    f"Files: {USER_SETTINGS_FILE.name}, {USER_SECRETS_FILE.name}."
                )
            except Exception as exc:  # noqa: BLE001
                st.error(f"Could not save settings: {exc}")

        if submitted_test:
            # Apply (but don't persist) so test_ai_provider sees the
            # current form values via os.environ.
            apply_user_settings_to_env(new_settings, new_secrets)
            with st.spinner(f"Testing {provider_choice}…"):
                model_for_test = {
                    "openai": new_settings.openai_model,
                    "anthropic": new_settings.anthropic_model,
                    "openrouter": new_settings.openrouter_model,
                }.get(provider_choice, "")
                key_for_test = provider_key_lookup.get(provider_choice, "")
                ok, msg = test_ai_provider(
                    provider_choice,
                    model=model_for_test,
                    api_key=key_for_test,
                )
            if ok:
                st.success(f"Test passed — {msg}")
            else:
                st.error(f"Test failed — {msg}")

    st.divider()

    # --- Execution mode (Phase 2b) ---------------------------------------
    _render_execution_mode_section(saved_settings, saved_secrets)

    st.divider()

    # --- Saved values summary --------------------------------------------
    st.subheader("Currently saved")
    st.dataframe(
        [
            {"setting": "Market",                 "value": saved_settings.market or "(US default)"},
            {"setting": "AI Provider",            "value": saved_settings.ai_provider or "(empty)"},
            {"setting": "Execution mode",         "value": saved_settings.agent_execution_mode or "(cli)"},
            {"setting": "Agent base URL",         "value": saved_settings.agent_base_url or "(127.0.0.1:9444)"},
            {"setting": "Agent API token",        "value": mask_key(saved_secrets.agent_api_token)},
            {"setting": "OpenAI model",           "value": saved_settings.openai_model or "(default)"},
            {"setting": "OpenAI API key",         "value": mask_key(saved_secrets.openai_api_key)},
            {"setting": "Anthropic model",        "value": saved_settings.anthropic_model or "(default)"},
            {"setting": "Anthropic API key",      "value": mask_key(saved_secrets.anthropic_api_key)},
            {"setting": "OpenRouter model",       "value": saved_settings.openrouter_model or "(auto)"},
            {"setting": "OpenRouter API key",     "value": mask_key(saved_secrets.openrouter_api_key)},
            {"setting": "OpenRouter site URL",    "value": saved_settings.openrouter_site_url or "(none)"},
            {"setting": "OpenRouter app name",    "value": saved_settings.openrouter_app_name or "(none)"},
        ],
        use_container_width=True,
        hide_index=True,
    )
    st.caption(
        f"Files: ``{USER_SETTINGS_FILE}`` + ``{USER_SECRETS_FILE}``. "
        "Secrets file is excluded from git and from the alpha ZIP. "
        "Keys are masked above (last 4 chars only)."
    )


def page_dashboard() -> None:
    st.title("Dashboard")

    # --- Chrome / Flow connectivity ---
    st.subheader("Connectivity")
    ok, info = chrome_reachable()
    if ok:
        st.success(f"Chrome reachable via {chrome_cdp_url()}")
        with st.expander("/json/version response"):
            st.code(info, language="json")
    else:
        st.error(f"Cannot reach Chrome CDP at {chrome_cdp_url()}.")
        st.caption(info)
        st.info(
            "On the Windows host, run `scripts\\start_chrome_debug.ps1` "
            "to start the debug Chrome profile, then refresh this page."
        )

    # --- Batch status ---
    total, counts = csv_status_counts()
    st.subheader(f"Current batch — {total} row(s)")
    if total == 0:
        st.info(
            "No products.csv yet. Go to **Manifest Builder**, save your "
            "prompts, then **Batch Controls → Load Manifest Fresh**."
        )
    else:
        labels = [
            ("pending", STATUS_PENDING),
            ("image_submitted", STATUS_IMAGE_SUBMITTED),
            ("image_approved", STATUS_IMAGE_APPROVED),
            ("image_rejected", STATUS_IMAGE_REJECTED),
            ("video_pending", STATUS_VIDEO_PENDING),
            ("video_submitted", STATUS_VIDEO_SUBMITTED),
            ("done", STATUS_DONE),
        ]
        cols = st.columns(len(labels))
        for col, (label, key) in zip(cols, labels):
            col.metric(label, counts.get(key, 0))

    # --- Next recommended action ---
    st.subheader("Next action")
    if not ok:
        st.warning("Start the host Chrome debug profile first.")
    elif total == 0:
        st.info("Author a manifest and Load Manifest Fresh.")
    elif counts.get(STATUS_PENDING, 0) > 0:
        st.info(
            f"{counts[STATUS_PENDING]} pending — run **Batch Controls → "
            f"Generate Images**."
        )
    elif counts.get(STATUS_IMAGE_SUBMITTED, 0) > 0:
        st.info(
            f"{counts[STATUS_IMAGE_SUBMITTED]} submitted — heart the good "
            f"ones in Flow Labs (your host Chrome window), then run "
            f"**Batch Controls → Sync Favorites**."
        )
    elif counts.get(STATUS_IMAGE_APPROVED, 0) > 0:
        st.info(
            f"{counts[STATUS_IMAGE_APPROVED]} approved — run **Batch "
            f"Controls → Generate Videos**."
        )
    elif counts.get(STATUS_VIDEO_SUBMITTED, 0) > 0:
        st.success("All rows have moved to video_submitted — batch done.")
    else:
        st.write("All rows have moved past the standard pipeline.")


def page_reference_images() -> None:
    st.title("Reference Images")
    st.caption(f"Folder: `{REFERENCE_IMAGES_DIR}`")
    REFERENCE_IMAGES_DIR.mkdir(parents=True, exist_ok=True)

    st.subheader("Upload")
    uploaded = st.file_uploader(
        "Drop product reference images (.jpg, .jpeg, .png, .webp)",
        accept_multiple_files=True,
        type=["jpg", "jpeg", "png", "webp"],
    )
    if uploaded:
        for f in uploaded:
            target = REFERENCE_IMAGES_DIR / f.name
            target.write_bytes(f.getvalue())
        st.success(f"Saved {len(uploaded)} file(s).")

    st.subheader("Current reference images")
    files = list_reference_images()
    if not files:
        st.info("No reference images yet.")
        return

    cols_per_row = 4
    for i, f in enumerate(files):
        if i % cols_per_row == 0:
            row = st.columns(cols_per_row)
        with row[i % cols_per_row]:
            st.image(str(f), caption=f.name, use_column_width=True)
            if st.button("Delete", key=f"del_{f.name}"):
                f.unlink()
                st.rerun()

    st.divider()
    st.subheader("Clear all reference images")
    confirm = st.checkbox(
        "I understand this permanently deletes every file in "
        f"{REFERENCE_IMAGES_DIR}.",
        key="clear_refs_confirm",
    )
    if confirm and st.button("Clear all", type="primary"):
        for f in files:
            try:
                f.unlink()
            except OSError as exc:
                st.warning(f"Couldn't delete {f.name}: {exc}")
        st.success("Cleared.")
        st.rerun()


def page_manifest() -> None:
    st.title("Manifest Builder")
    st.caption(f"File: `{MANIFEST_PATH}`")
    MANIFEST_PATH.parent.mkdir(parents=True, exist_ok=True)
    if not MANIFEST_PATH.exists():
        MANIFEST_PATH.write_text("", encoding="utf-8")

    current = MANIFEST_PATH.read_text(encoding="utf-8")
    edited = st.text_area(
        "Manifest (Markdown)",
        value=current,
        height=420,
        key="manifest_editor",
    )

    c1, c2, c3 = st.columns(3)
    if c1.button("Save"):
        MANIFEST_PATH.write_text(edited, encoding="utf-8")
        st.success(f"Saved {len(edited)} chars to {MANIFEST_PATH}.")
    if c2.button("Re-load from disk"):
        st.rerun()
    if c3.button("Validate (CLI)"):
        run_cli(["--validate-manifest", str(MANIFEST_PATH)])

    st.subheader("Parsed products")
    try:
        entries = parse_manifest(MANIFEST_PATH)
    except FileNotFoundError as exc:
        st.error(str(exc))
        return
    except Exception as exc:  # noqa: BLE001
        st.error(f"Failed to parse manifest: {exc}")
        return

    if not entries:
        st.info(
            "No sections detected. Each product must start with a line "
            "like `## 01` or `### Product 01`."
        )
        return

    rows = []
    issues = 0
    for e in entries:
        resolved = (
            resolve_reference_image(e.reference_image, SETTINGS)
            if e.reference_image
            else None
        )
        ref_exists = bool(resolved and resolved.exists())
        missing: list[str] = []
        if not e.product_name:
            missing.append("Product Name")
        if not e.reference_image:
            missing.append("Reference Image")
        if not e.image_prompt:
            missing.append("Image Prompt")
        if missing or (e.reference_image and not ref_exists):
            issues += 1
        rows.append(
            {
                "id": e.id,
                "product_name": e.product_name or "(missing)",
                "reference_image": e.reference_image or "(missing)",
                "resolved": str(resolved.relative_to(REPO_ROOT)) if resolved else "",
                "exists": ref_exists,
                "image_prompt_chars": len(e.image_prompt),
                "video_prompt_chars": len(e.video_prompt),
                "status": e.status,
                "missing": ", ".join(missing) if missing else "",
            }
        )

    st.dataframe(rows, use_container_width=True)
    if issues:
        st.warning(f"{issues} section(s) have missing fields or unresolved images.")
    else:
        st.success("All sections look good.")


def page_batch() -> None:
    st.title("Batch Controls")
    st.caption(
        "Each button subprocesses `python main.py` — the same code path "
        "the CLI uses. Don't navigate away while a batch is running."
    )

    default_limit = st.session_state.get("batch_limit", 30)
    limit = st.number_input(
        "Limit (rows per batch)",
        min_value=1, max_value=200, value=default_limit, step=1,
    )
    st.session_state["batch_limit"] = limit

    st.subheader("Quick checks")
    qc1, qc2, qc3 = st.columns(3)
    if qc1.button("Check Browser"):
        run_cli(["--check-browser"])
    if qc2.button("List Status"):
        run_cli(["--list-status"])
    if qc3.button("Validate Manifest"):
        run_cli(["--validate-manifest", str(MANIFEST_PATH)])

    st.divider()
    st.subheader("Load manifest (destructive when --fresh)")
    fresh = st.checkbox(
        "Use --fresh (backs up the current CSV and starts a new batch).",
        value=False,
        key="load_fresh_checkbox",
    )
    if fresh:
        st.warning(
            "This will move the existing `inputs/products.csv` aside to "
            "`products.csv.bak.<timestamp>` and rewrite it from the "
            "current manifest. Already-generated rows from the old batch "
            "are no longer tracked in the CSV."
        )
    if st.button("Load Manifest" + (" Fresh" if fresh else "")):
        args = ["--load-manifest", str(MANIFEST_PATH)]
        if fresh:
            args.append("--fresh")
        run_cli(args)

    st.divider()
    st.subheader("Generate")
    g1, g2, g3 = st.columns(3)
    if g1.button(f"Generate Images (limit {limit})"):
        run_cli(["--generate-images", "--limit", str(limit)])
    if g2.button("Sync Favorites"):
        run_cli(["--sync-favorites"])
    if g3.button(f"Generate Videos (limit {limit})"):
        run_cli(["--generate-videos", "--limit", str(limit)])


def page_logs() -> None:
    st.title("Logs")

    st.subheader("Today's log")
    today = LOGS_DIR / f"{datetime.now():%Y-%m-%d}.log"
    if today.exists():
        text = today.read_text(encoding="utf-8", errors="replace")
        st.code(text[-20000:], language="text")
        st.caption(
            f"{today.name} — {today.stat().st_size:,} bytes (showing last 20 000)."
        )
    else:
        st.info(f"No log file at {today} yet.")

    st.subheader("Last command output (this session)")
    last_args = st.session_state.get("last_command_args")
    last_out = st.session_state.get("last_command_output")
    if last_args and last_out:
        st.caption(f"`python main.py {' '.join(last_args)}`")
        st.code(last_out[-20000:], language="text")
    else:
        st.info("No commands run from the UI yet.")

    st.subheader("Recent error screenshots")
    screenshots: list[Path] = []
    if LOGS_DIR.exists():
        screenshots = sorted(
            LOGS_DIR.glob("error_*.png"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )[:8]
    if screenshots:
        cols = st.columns(2)
        for i, img in enumerate(screenshots):
            with cols[i % 2]:
                st.image(str(img), caption=img.name, use_column_width=True)
    else:
        st.info("No error_*.png files in outputs/logs/.")

    st.subheader("All log files")
    if LOGS_DIR.exists():
        log_files = sorted(
            LOGS_DIR.glob("*.log"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )[:20]
        for f in log_files:
            st.text(f"{f.name}  ({f.stat().st_size:,} bytes)")
    else:
        st.info(f"No {LOGS_DIR} directory yet.")


def page_settings() -> None:
    st.title("Settings")

    st.subheader("Per-session UI defaults")
    limit = st.number_input(
        "Default limit for Generate Images / Videos",
        min_value=1, max_value=200,
        value=st.session_state.get("batch_limit", 30),
        step=1,
    )
    st.session_state["batch_limit"] = limit
    st.caption("Stored in this Streamlit session only — not written to .env.")

    st.subheader("Environment (read-only)")
    keys = [
        "BROWSER_MODE", "CHROME_CDP_URL", "FLOW_LABS_URL",
        "HEADLESS", "GENERATION_TIMEOUT_SECONDS", "SELECTOR_TIMEOUT_MS",
        "SLOW_MO_MS", "SAVE_OUTPUT_IMAGE", "VERIFY_GENERATION_STARTED",
        "CAPTURE_TIMEOUT_SECONDS",
    ]
    st.dataframe(
        [{"name": k, "value": os.environ.get(k, "(unset)")} for k in keys],
        use_container_width=True,
    )
    st.caption(
        "Values come from docker-compose.yml. BROWSER_MODE and "
        "CHROME_CDP_URL are hardcoded for Docker safety; the rest fall "
        "through from your host `.env` via Compose's variable "
        "substitution. Restart the UI service to pick up changes:\n\n"
        "    docker compose restart ui"
    )

    st.subheader("Project paths")
    paths = {
        "Repo root": REPO_ROOT,
        "Products CSV": PRODUCTS_CSV,
        "Reference Images": REFERENCE_IMAGES_DIR,
        "Manifest": MANIFEST_PATH,
        "Logs": LOGS_DIR,
    }
    st.dataframe(
        [
            {
                "name": k,
                "path": str(v),
                "exists": Path(v).exists(),
            }
            for k, v in paths.items()
        ],
        use_container_width=True,
    )


# ---------------------------------------------------------------------------
# App entrypoint
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# AI Product Intake page
# ---------------------------------------------------------------------------


def _current_batch() -> str | None:
    return st.session_state.get("current_batch")


def _provider_status_chip(provider) -> None:
    ok, msg = provider.is_configured()
    if ok:
        st.success(f"Provider **{provider.name}** ready ({msg}).")
    else:
        st.error(f"Provider **{provider.name}** is NOT configured: {msg}")


def _provider_model_input(provider_name: str) -> None:
    """Show a model text input for the selected provider and persist it
    to os.environ so the provider class picks it up."""
    if provider_name == "openai":
        env_key, default = "OPENAI_MODEL", "gpt-4o-mini"
    elif provider_name == "anthropic":
        env_key, default = "ANTHROPIC_MODEL", "claude-3-5-sonnet-latest"
    elif provider_name == "openrouter":
        env_key, default = "OPENROUTER_MODEL", "anthropic/claude-3.5-sonnet"
    else:
        return
    current = os.environ.get(env_key, default)
    new_val = st.text_input(f"{env_key}", value=current, key=f"model_input_{env_key}")
    os.environ[env_key] = new_val


def _ai_provider_panel() -> None:
    """Compact provider picker for the AI Prompts / intake pages.

    Reads the current provider + model from os.environ (populated from
    settings.local.json on startup). For full editing (keys, model
    selection per provider, test button) the user is pointed at the
    dedicated Setup page.
    """
    with st.expander("AI provider", expanded=False):
        current_provider = (os.environ.get("AI_PROVIDER") or "manual").strip().lower()
        try:
            provider = get_provider(current_provider)
            _provider_status_chip(provider)
        except Exception as exc:  # noqa: BLE001
            st.error(f"Could not load provider {current_provider}: {exc}")
        st.caption(
            "To change provider, model, or API key, open **Setup** from "
            "the sidebar. Changes save to data/settings.local.json + "
            "data/secrets.local.json and apply immediately."
        )


def _ensure_session_field(key: str, default: str) -> None:
    """Initialize session_state[key] from default only if absent."""
    if key not in st.session_state:
        st.session_state[key] = default


def _render_product_card(
    batch_id: str, product: ProductCard, all_products: list[ProductCard]
) -> None:
    """Per-card expander body: edit fields, upload refs, generate AI."""
    pid = product.id

    # Initialize session_state keys from disk on first render so widget
    # values aren't lost across reruns.
    for fname, val in {
        "name":  product.product_name,
        "url":   product.tiktok_url,
        "desc":  product.product_description,
        "notes": product.notes,
        "img":   product.image_prompt,
        "vid":   product.video_prompt,
        "hook":  product.hook,
        "cap":   product.caption,
        "cat":   product.category,
        "store": product.store_environment,
        "place": product.placement_type,
    }.items():
        _ensure_session_field(f"pf_{fname}_{pid}", val)

    # Show the active prompt mode + the selected retailer (if any) at
    # the top of the card so the user can tell at a glance which store
    # the next AI run will target.
    if _active_market() == "UK":
        selected = (product.store_environment or "").strip()
        if selected:
            st.caption(f"🇬🇧 UK Retail Store Display · **Retailer:** {selected}")
        else:
            st.caption("🇬🇧 UK Retail Store Display · **Retailer:** (let AI choose)")

    col_a, col_b = st.columns([2, 1])
    with col_a:
        st.text_input("Product Name", key=f"pf_name_{pid}")
        st.text_input("TikTok URL", key=f"pf_url_{pid}")
        st.text_area("Product Description", key=f"pf_desc_{pid}", height=80)
        st.text_area("Notes", key=f"pf_notes_{pid}", height=60)
    with col_b:
        st.markdown("**Reference images (1–3)**")
        uploaded = st.file_uploader(
            "Upload",
            accept_multiple_files=True,
            type=["jpg", "jpeg", "png", "webp"],
            key=f"pf_upload_{pid}",
            label_visibility="collapsed",
        )
        if uploaded:
            signature = tuple((f.name, f.size) for f in uploaded[:3])
            sig_key = f"pf_uploadsig_{pid}"
            if st.session_state.get(sig_key) != signature:
                st.session_state[sig_key] = signature
                product.reference_images = []
                for i, f in enumerate(uploaded[:3]):
                    role = REFERENCE_ROLES[i]
                    ext = Path(f.name).suffix or ".jpg"
                    saved = save_reference_image(
                        batch_id, pid, role, f.getvalue(), ext
                    )
                    product.reference_images.append(reference_image_rel(saved))
                save_batch_products(batch_id, all_products)
                st.success(f"Saved {len(uploaded[:3])} reference image(s).")

        if product.reference_images:
            thumb_cols = st.columns(len(product.reference_images))
            for col, ref in zip(thumb_cols, product.reference_images):
                full = (Path(ref) if Path(ref).is_absolute() else (ROOT / ref))
                if full.exists():
                    col.image(str(full), caption=Path(ref).name, use_column_width=True)

        # Clipboard paste per-card. Auto-fills the next free role; refuses
        # to overwrite once all 3 slots are taken.
        if HAS_PASTE_BUTTON:
            next_slot = product.next_available_role()
            paste_label = (
                f"📋 Paste image (→ {next_slot})"
                if next_slot
                else "All 3 reference slots full — clear one to paste"
            )
            paste_result = _paste_image_button(
                paste_label,
                key=f"pf_paste_{pid}",
                errors="ignore",
            )
            if (
                paste_result
                and getattr(paste_result, "image_data", None) is not None
                and next_slot
            ):
                ok, msg = _save_pasted_image_for_product(
                    batch_id, product, all_products, paste_result.image_data
                )
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    st.divider()

    # AI generation
    provider_name = os.environ.get("AI_PROVIDER", "manual")
    btn_label = f"Generate AI prompts ({provider_name})"
    if st.button(btn_label, key=f"pf_gen_{pid}"):
        # Pull latest field values from session_state before calling.
        product.product_name        = st.session_state[f"pf_name_{pid}"]
        product.tiktok_url          = st.session_state[f"pf_url_{pid}"]
        product.product_description = st.session_state[f"pf_desc_{pid}"]
        product.notes               = st.session_state[f"pf_notes_{pid}"]
        save_batch_products(batch_id, all_products)
        _call_ai_for_product(batch_id, product, all_products, provider_name)

    # AI-output fields (editable). Session state holds whatever the user
    # last typed OR the value set after AI generation.
    st.text_area("Image Prompt", key=f"pf_img_{pid}", height=200)
    st.text_area("Video Prompt", key=f"pf_vid_{pid}", height=140)
    cc1, cc2 = st.columns(2)
    cc1.text_area("Hook", key=f"pf_hook_{pid}", height=80)
    cc2.text_input("Caption", key=f"pf_cap_{pid}")
    cc3, cc4, cc5 = st.columns(3)
    cc3.text_input("Category",          key=f"pf_cat_{pid}")
    # When MARKET=UK, "Store Environment" becomes a curated dropdown of
    # the 14 UK retailers covered by the prompt library — saves the
    # user from typing exact spelling and serves as a manual override
    # of whatever the AI picked. Empty option = "let the AI choose".
    if _active_market() == "UK":
        current_store = (st.session_state.get(f"pf_store_{pid}") or "").strip()
        options = ["(let AI choose)", *UK_RETAILERS]
        try:
            idx = options.index(current_store) if current_store in options else 0
        except ValueError:
            idx = 0
        picked = cc4.selectbox(
            "UK Retailer (override)",
            options,
            index=idx,
            key=f"pf_store_select_{pid}",
            help=(
                "Pick a UK retailer to force the AI prompt onto a "
                "specific store, or leave on '(let AI choose)' to let "
                "the model pick based on the product category."
            ),
        )
        # Mirror selection into the underlying free-text key so the
        # save handler (which reads from pf_store_{pid}) sees the right
        # value with no other code changes.
        st.session_state[f"pf_store_{pid}"] = "" if picked.startswith("(") else picked
    else:
        cc4.text_input("Store Environment", key=f"pf_store_{pid}")
    cc5.text_input("Placement Type",    key=f"pf_place_{pid}")

    if product.warnings:
        st.warning("Warnings: " + "; ".join(product.warnings))

    st.divider()
    save_col, del_col = st.columns([4, 1])
    if save_col.button("Save product", key=f"pf_save_{pid}"):
        # Pull all fields from session_state and persist.
        product.product_name        = st.session_state[f"pf_name_{pid}"]
        product.tiktok_url          = st.session_state[f"pf_url_{pid}"]
        product.product_description = st.session_state[f"pf_desc_{pid}"]
        product.notes               = st.session_state[f"pf_notes_{pid}"]
        product.image_prompt        = st.session_state[f"pf_img_{pid}"]
        product.video_prompt        = st.session_state[f"pf_vid_{pid}"]
        product.hook                = st.session_state[f"pf_hook_{pid}"]
        product.caption             = st.session_state[f"pf_cap_{pid}"]
        product.category            = st.session_state[f"pf_cat_{pid}"]
        product.store_environment   = st.session_state[f"pf_store_{pid}"]
        product.placement_type      = st.session_state[f"pf_place_{pid}"]
        # Under strict blanket-video-prompt mode, image_prompt alone is
        # enough to be READY. is_ready_to_export() encodes that gate
        # (name + reference + image_prompt).
        product.status = (
            PRODUCT_STATUS_READY if product.is_ready_to_export() else PRODUCT_STATUS_DRAFT
        )
        save_batch_products(batch_id, all_products)
        st.success(f"Saved (status={product.status}).")
    if del_col.button("Delete", key=f"pf_del_{pid}", type="primary"):
        remaining = [p for p in all_products if p.id != pid]
        save_batch_products(batch_id, remaining)
        # Clear the card's session_state so it doesn't leak.
        for k in list(st.session_state.keys()):
            if k.endswith(f"_{pid}"):
                del st.session_state[k]
        st.rerun()


def _call_ai_for_product(
    batch_id: str,
    product: ProductCard,
    all_products: list[ProductCard],
    provider_name: str,
) -> None:
    provider = get_provider(provider_name)
    ok, msg = provider.is_configured()
    if not ok:
        st.error(f"AI provider not configured: {msg}")
        return

    product_dict = {
        "product_name":         product.product_name,
        "tiktok_url":           product.tiktok_url,
        "product_description":  product.product_description,
        "notes":                product.notes,
        "reference_filenames":  [Path(r).name for r in product.reference_images],
        "category_hint":        product.category,
        "store_hint":           product.store_environment,
        "placement_hint":       product.placement_type,
    }
    raw_text: str | None = None
    try:
        with st.spinner(f"Generating prompts via {provider_name}..."):
            output = provider.generate_product_prompts(product_dict)
    except Exception as exc:  # noqa: BLE001
        # Try to surface the raw text the provider received, if any.
        raw_text = getattr(exc, "args", [None])[0]
        st.error(f"AI generation failed: {exc}")
        if isinstance(raw_text, str):
            # NOTE: this runs inside the per-product expander, and
            # Streamlit forbids nesting expanders. Use a checkbox-gated
            # code block instead.
            if st.checkbox(
                "Show raw response from model",
                key=f"pf_show_raw_err_{product.id}",
            ):
                st.code(raw_text)
        return

    ok_v, problems = validate_ai_output(output)
    if not ok_v:
        st.warning("AI output had issues: " + "; ".join(problems))
        if st.checkbox(
            "Show raw AI output JSON",
            key=f"pf_show_raw_ok_{product.id}",
        ):
            st.json(output)

    # Apply to product. Push into session_state so the widgets refresh.
    pid = product.id
    mapping = {
        "image_prompt":      "pf_img_",
        "video_prompt":      "pf_vid_",
        "hook":              "pf_hook_",
        "caption":           "pf_cap_",
        "category":          "pf_cat_",
        "store_environment": "pf_store_",
        "placement_type":    "pf_place_",
    }
    for field_name, key_prefix in mapping.items():
        new_value = (output.get(field_name) or "").strip() if isinstance(output.get(field_name), str) else ""
        setattr(product, field_name, new_value)
        st.session_state[f"{key_prefix}{pid}"] = new_value

    raw_warnings = output.get("warnings") or []
    if isinstance(raw_warnings, list):
        product.warnings = [str(w) for w in raw_warnings]
    else:
        product.warnings = [str(raw_warnings)]

    product.status = (
        PRODUCT_STATUS_READY if product.has_prompts() else PRODUCT_STATUS_DRAFT
    )
    save_batch_products(batch_id, all_products)
    st.success("Prompts generated. Review and Save.")
    st.rerun()


def _status_chip(label: str, ok: bool) -> str:
    """Compact green/red chip used inside each card header."""
    return f"{'✅' if ok else '⬜'} {label}"


def _save_pasted_image_for_product(
    batch_id: str,
    product: ProductCard,
    all_products: list[ProductCard],
    image_obj,
) -> tuple[bool, str]:
    """Save a PIL.Image into the next free reference slot for `product`.

    Returns (ok, message). Refuses to overwrite filled slots — if all
    three are taken, returns (False, "all 3 slots filled").
    """
    role = product.next_available_role()
    if role is None:
        return False, "All 3 reference slots are filled; clear one before pasting."
    # Always serialize as PNG for consistency (pasted images often have
    # no extension info).
    from io import BytesIO
    buf = BytesIO()
    try:
        image_obj.save(buf, format="PNG")
    except Exception as exc:  # noqa: BLE001
        return False, f"Could not encode pasted image: {exc}"
    saved = save_reference_image(batch_id, product.id, role, buf.getvalue(), "png")
    rel = reference_image_rel(saved)
    # Replace existing slot (defensive) or append.
    refs: list[str] = []
    found_role = False
    for r in product.reference_images:
        if f"_{role}." in Path(r).name:
            refs.append(rel)
            found_role = True
        else:
            refs.append(r)
    if not found_role:
        refs.append(rel)
    product.reference_images = refs
    save_batch_products(batch_id, all_products)
    return True, f"Saved pasted image as {role} ({Path(rel).name})."


# ---------------------------------------------------------------------------
# Bulk add — paste JSON array of {title, url}
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# Kalodata Excel import — reads LIST_PRODUCT sheet, downloads product
# images, creates one card per row.
# ---------------------------------------------------------------------------


_KALODATA_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def _kalodata_str(v) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "none") else s


def _kalodata_get(row: dict, *candidates: str) -> str:
    """Read a value from a Kalodata row tolerating column-name variants."""
    for c in candidates:
        if c in row:
            value = _kalodata_str(row[c])
            if value:
                return value
    return ""


# Canonical field name -> ordered list of column aliases seen in Kalodata
# exports. Kalodata renames columns between report types ("Best Sellers",
# "Favorites", "Focus") — every new variant goes here, NOT scattered
# through the call sites.
KALODATA_FIELD_ALIASES: dict[str, list[str]] = {
    # Core
    "product_name":   ["Product Name"],
    "img_url":        ["img_url", "Image Link", "Image URL", "ImageUrl"],
    "category":       ["Category"],
    "tiktok_url":     ["TikTokUrl", "TikTok URL", "TikTok Link", "TiktokUrl"],
    "kalodata_url":   ["KalodataUrl", "Kalodata URL", "Kalodata Details Link"],
    # Pricing / commerce metrics
    "price":          ["Avg. Unit Price($)", "Price($)", "Price"],
    "price_range":    ["Price Range($)"],
    "commission":     ["Commission Rate"],
    "revenue":        ["Revenue($)", "Revenue"],
    "revenue_growth": ["Revenue Growth Rate"],
    "live_revenue":   ["Live Revenue($)"],
    "video_revenue":  ["Video Revenue($)"],
    "card_revenue":   ["Product Card Revenue"],
    # Creator / engagement metrics
    "creators":       ["Creator Count", "Creator Number"],
    "new_creators":   ["New Creator Count"],
    "conversion":     ["Creator Conversion Ratio"],
    "video_count":    ["Video Count"],
    "new_videos":     ["New Video Count"],
    # Product / report metadata
    "rating":         ["Product Rating"],
    "item_sold":      ["Item Sold", "Items Sold"],
    "item_sold_growth": ["Item Sold Growth Rate"],
    "launch_date":    ["Launch Date"],
    "date_range":     ["Date Range"],
    "remarks":        ["Remarks"],
}


def _kalodata_field(row: dict, canonical: str) -> str:
    """Read a canonical Kalodata field, trying every known alias."""
    aliases = KALODATA_FIELD_ALIASES.get(canonical) or [canonical]
    return _kalodata_get(row, *aliases)


_PREFERRED_KALODATA_SHEETS = ["LIST_PRODUCT", "LIST_PRODUCT_FOCUS"]
_META_SHEET_NAMES = {"intro", "info", "about", "cover", "metadata", "summary"}


def _select_kalodata_sheet(sheet_names: list[str]) -> str:
    """Pick the most likely product-list sheet from a Kalodata workbook.

    Preference order:
        1. Exact match for any known preferred sheet (LIST_PRODUCT,
           LIST_PRODUCT_FOCUS).
        2. Any sheet whose name starts with LIST_PRODUCT (handles future
           Kalodata variants like LIST_PRODUCT_FAVORITES).
        3. The first non-metadata sheet.
    """
    upper_to_actual = {n.upper(): n for n in sheet_names}
    for name in _PREFERRED_KALODATA_SHEETS:
        if name in sheet_names:
            return name
        if name.upper() in upper_to_actual:
            return upper_to_actual[name.upper()]
    for name in sheet_names:
        if name.upper().startswith("LIST_PRODUCT"):
            return name
    for name in sheet_names:
        if name.strip().lower() not in _META_SHEET_NAMES:
            return name
    raise ValueError(
        f"No product sheet found. Sheets present: "
        f"{', '.join(sheet_names) or '(none)'}"
    )


def _import_from_kalodata_excel(file_bytes: bytes) -> tuple[str, list[dict]]:
    """Parse a Kalodata .xlsx export.

    Returns ``(sheet_name_used, rows)``. ``rows`` is a list of dicts
    keyed on the exact column headers Kalodata wrote. Callers should
    use :func:`_kalodata_field` to read by canonical name so column
    renames don't propagate everywhere.
    """
    from io import BytesIO
    import openpyxl

    wb = openpyxl.load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet_name = _select_kalodata_sheet(list(wb.sheetnames))
    sheet = wb[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers_raw = next(rows_iter)
    except StopIteration:
        return sheet_name, []
    headers = [str(h).strip() if h is not None else "" for h in headers_raw]
    out: list[dict] = []
    for row_values in rows_iter:
        if all(v in (None, "") for v in row_values):
            continue
        record: dict = {}
        for h, v in zip(headers, row_values):
            if h:
                record[h] = v
        out.append(record)
    return sheet_name, out


def _download_image(url: str, timeout: int = 20) -> tuple[bytes, str]:
    """Fetch an image URL, return (bytes, extension_without_dot).

    Infers the extension from Content-Type, then from URL suffix, then
    from magic bytes. Defaults to ``jpg`` if nothing fits. Sends a
    browser-like User-Agent so CDNs that block unknown clients (TikTok)
    still serve the bytes.
    """
    from urllib.parse import urlparse
    import urllib.request

    if not url.startswith(("http://", "https://")):
        raise ValueError(f"unsupported URL scheme: {url[:60]!r}")

    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": _KALODATA_USER_AGENT,
            "Accept": "image/avif,image/webp,image/apng,image/svg+xml,image/*,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        ctype = (resp.headers.get("Content-Type", "") or "").lower()
        data = resp.read()
    if not data:
        raise ValueError("empty response body")

    ext = None
    if "jpeg" in ctype or "jpg" in ctype:
        ext = "jpg"
    elif "png" in ctype:
        ext = "png"
    elif "webp" in ctype:
        ext = "webp"
    elif "gif" in ctype:
        ext = "gif"

    if not ext:
        path = urlparse(url).path.lower()
        for candidate in ("jpg", "jpeg", "png", "webp"):
            if path.endswith(f".{candidate}"):
                ext = candidate
                break

    if not ext:
        # Magic bytes
        if data[:3] == b"\xff\xd8\xff":
            ext = "jpg"
        elif data[:8] == b"\x89PNG\r\n\x1a\n":
            ext = "png"
        elif len(data) >= 12 and data[:4] == b"RIFF" and data[8:12] == b"WEBP":
            ext = "webp"
        else:
            ext = "jpg"  # last-resort default; TikTok CDN often misreports type

    if ext == "jpeg":
        ext = "jpg"
    return data, ext


def _create_cards_from_kalodata(
    batch_id: str,
    all_products: list[ProductCard],
    rows: list[dict],
) -> None:
    """Build one ProductCard per row, downloading images best-effort.

    Saves the batch after the whole run so a single failure doesn't
    abandon work-in-progress. A failed image becomes a card-level
    warning; the card is still created so the user can fix it via the
    Image Pool or clipboard paste.
    """
    if not rows:
        return
    progress = st.progress(0.0)
    status_box = st.empty()
    log_lines: list[str] = []
    log_box = st.empty()

    created = 0
    image_ok = 0
    image_failed = 0

    for i, row in enumerate(rows, start=1):
        name = _kalodata_field(row, "product_name") or "(unnamed)"
        status_box.text(f"[{i}/{len(rows)}] {name[:60]}")

        pid = new_product_id()
        category   = _kalodata_field(row, "category")
        tiktok_url = _kalodata_field(row, "tiktok_url")
        kalo_url   = _kalodata_field(row, "kalodata_url")
        img_url    = _kalodata_field(row, "img_url")

        # Notes preserve every Kalodata metric the prompt + downstream
        # tooling might want. Empty fields are omitted so the note stays
        # tight regardless of which Kalodata report type was exported.
        note_fields: list[tuple[str, str]] = [
            ("Date Range",        "date_range"),
            ("Launch Date",       "launch_date"),
            ("Rating",            "rating"),
            ("Items Sold",        "item_sold"),
            ("Items Sold Growth", "item_sold_growth"),
            ("Avg Unit Price",    "price"),
            ("Price Range",       "price_range"),
            ("Commission",        "commission"),
            ("Revenue",           "revenue"),
            ("Revenue Growth",    "revenue_growth"),
            ("Live Revenue",      "live_revenue"),
            ("Video Revenue",     "video_revenue"),
            ("Card Revenue",      "card_revenue"),
            ("Creators",          "creators"),
            ("New Creators",      "new_creators"),
            ("Conversion",        "conversion"),
            ("Video Count",       "video_count"),
            ("New Videos",        "new_videos"),
            ("Remarks",           "remarks"),
            ("Kalodata URL",      "kalodata_url"),
        ]
        note_parts: list[str] = []
        for label, canonical in note_fields:
            v = _kalodata_field(row, canonical)
            if v:
                note_parts.append(f"{label}: {v}")
        notes = " | ".join(note_parts)

        description_bits = [name]
        if category:
            description_bits.append(f"[{category}]")
        product_description = " ".join(description_bits)

        card = ProductCard(
            id=pid,
            product_name=name if name != "(unnamed)" else "",
            original_title=name,
            tiktok_url=tiktok_url,
            product_description=product_description,
            notes=notes,
            category=category,
            status=PRODUCT_STATUS_DRAFT,
        )

        if img_url:
            try:
                data, ext = _download_image(img_url)
                saved = save_reference_image(batch_id, pid, "primary", data, ext)
                card.reference_images = [reference_image_rel(saved)]
                image_ok += 1
                log_lines.append(f"  OK   {name[:60]} ({len(data) // 1024} KB .{ext})")
            except Exception as exc:  # noqa: BLE001
                image_failed += 1
                card.warnings.append(f"image download failed: {exc}")
                log_lines.append(f"  IMG-FAIL {name[:60]}: {exc}")
        else:
            image_failed += 1
            card.warnings.append("no img_url in Kalodata row")
            log_lines.append(f"  NO-IMG {name[:60]}")

        all_products.append(card)
        created += 1
        progress.progress(i / len(rows))
        log_box.code("\n".join(log_lines[-25:]), language="text")

    save_batch_products(batch_id, all_products)
    status_box.empty()
    progress.empty()

    summary = (
        f"Created {created} product card(s). "
        f"Images: {image_ok} ok, {image_failed} need manual attach."
    )
    if image_failed:
        st.warning(summary)
    else:
        st.success(summary)


def _kalodata_import_section(
    batch_id: str, products: list[ProductCard]
) -> None:
    with st.expander("Import from Kalodata Excel", expanded=False):
        st.caption(
            "Upload a Kalodata .xlsx export. Reads the `LIST_PRODUCT` "
            "sheet, downloads each row's image, and creates a product "
            "card per row with metrics in notes."
        )
        uploaded = st.file_uploader(
            "Kalodata .xlsx export",
            type=["xlsx"],
            key=f"kalodata_uploader_{batch_id}",
        )
        if uploaded:
            sig = (uploaded.name, uploaded.size)
            if st.session_state.get("kalodata_sig") != sig:
                try:
                    sheet_name, rows = _import_from_kalodata_excel(uploaded.getvalue())
                except Exception as exc:  # noqa: BLE001
                    st.error(f"Failed to read xlsx: {exc}")
                    st.session_state.pop("kalodata_rows", None)
                    st.session_state.pop("kalodata_sheet", None)
                    st.session_state.pop("kalodata_sig", None)
                    return
                st.session_state["kalodata_rows"] = rows
                st.session_state["kalodata_sheet"] = sheet_name
                st.session_state["kalodata_sig"] = sig

        rows: list[dict] = st.session_state.get("kalodata_rows", [])
        sheet_name = st.session_state.get("kalodata_sheet", "?")
        if not rows:
            return

        st.success(f"Loaded {len(rows)} row(s) from sheet `{sheet_name}`.")

        preview = []
        for r in rows:
            preview.append({
                "Product Name": _kalodata_field(r, "product_name")[:80],
                "Category":     _kalodata_field(r, "category"),
                "Price":        _kalodata_field(r, "price"),
                "Commission":   _kalodata_field(r, "commission"),
                "Revenue":      _kalodata_field(r, "revenue"),
                "Growth":       _kalodata_field(r, "revenue_growth"),
                "Creators":     _kalodata_field(r, "creators"),
                "TikTokUrl":    _kalodata_field(r, "tiktok_url")[:60],
                "img_url":      "yes" if _kalodata_field(r, "img_url") else "no",
            })
        st.dataframe(preview, use_container_width=True)

        cols = st.columns([2, 1])
        scope = cols[0].radio(
            "Import scope",
            ["All rows", "First N rows"],
            horizontal=True,
            key="kalodata_scope",
        )
        if scope == "First N rows":
            n = cols[1].number_input(
                "N",
                min_value=1,
                max_value=len(rows),
                value=min(20, len(rows)),
                step=1,
                key="kalodata_n",
            )
            target_rows = rows[: int(n)]
        else:
            target_rows = rows

        if st.button(
            f"Create {len(target_rows)} Product Cards from Kalodata",
            type="primary",
            key="kalodata_create",
        ):
            _create_cards_from_kalodata(batch_id, products, target_rows)
            st.session_state.pop("kalodata_rows", None)
            st.session_state.pop("kalodata_sheet", None)
            st.session_state.pop("kalodata_sig", None)
            st.rerun()


def _bulk_add_section(batch_id: str, products: list[ProductCard]) -> None:
    with st.expander("Bulk add products (paste JSON)", expanded=False):
        st.caption(
            "Paste a JSON array of `{title, url}` objects (e.g. exported "
            "from a Kalodata or TikTok scrape). Click Parse to preview, "
            "then Create Product Cards."
        )
        sample = (
            '[\n'
            '  {"title": "Cool Gadget - TikTok Shop", "url": "https://www.tiktok.com/shop/pdp/..."},\n'
            '  {"title": "Beauty Cream 2oz", "url": "https://www.kalodata.com/product/..."}\n'
            ']'
        )
        json_text = st.text_area(
            "JSON array",
            value=st.session_state.get("bulk_json_text", ""),
            placeholder=sample,
            height=180,
            key="bulk_json_text_input",
        )
        st.session_state["bulk_json_text"] = json_text

        c1, c2 = st.columns([1, 4])
        if c1.button("Parse Products"):
            try:
                parsed = json.loads(json_text)
            except json.JSONDecodeError as exc:
                st.error(f"Invalid JSON: {exc.msg} (line {exc.lineno}, col {exc.colno})")
                st.session_state.pop("bulk_parsed", None)
            else:
                if not isinstance(parsed, list):
                    st.error("JSON must be an array of objects.")
                    st.session_state.pop("bulk_parsed", None)
                else:
                    rows: list[dict] = []
                    skipped = 0
                    for i, item in enumerate(parsed):
                        if not isinstance(item, dict):
                            skipped += 1
                            continue
                        title = (item.get("title") or "").strip()
                        url   = (item.get("url") or "").strip()
                        if not title or not url:
                            skipped += 1
                            continue
                        rows.append(
                            {
                                "original_title": title,
                                "product_name": clean_product_title(title),
                                "tiktok_url": url,
                                "source": detect_url_source(url),
                                "has_image": False,
                            }
                        )
                    st.session_state["bulk_parsed"] = rows
                    msg = f"Parsed {len(rows)} product(s)."
                    if skipped:
                        msg += f" Skipped {skipped} (missing title or url, or not an object)."
                    st.success(msg)

        parsed_rows = st.session_state.get("bulk_parsed", [])
        if parsed_rows:
            st.dataframe(parsed_rows, use_container_width=True)
            if st.button("Create product cards", type="primary"):
                added = 0
                for row in parsed_rows:
                    card = ProductCard(
                        id=new_product_id(),
                        product_name=row["product_name"],
                        original_title=row["original_title"],
                        tiktok_url=row["tiktok_url"],
                        product_description=row["product_name"],
                        status=PRODUCT_STATUS_DRAFT,
                    )
                    products.append(card)
                    added += 1
                save_batch_products(batch_id, products)
                st.session_state.pop("bulk_parsed", None)
                st.session_state["bulk_json_text"] = ""
                st.success(f"Added {added} product card(s).")
                st.rerun()


# ---------------------------------------------------------------------------
# Image pool — bulk uploads + per-product attach buttons
# ---------------------------------------------------------------------------


def _image_pool_section(batch_id: str, products: list[ProductCard]) -> None:
    with st.expander("Image pool (bulk upload + attach)", expanded=False):
        pool_dir = image_pool_dir(batch_id)
        st.caption(
            f"Bulk-uploaded images live in `data/batches/{batch_id}/image_pool/`. "
            "Filenames don't have to match product names — you assign them "
            "visually below."
        )

        uploaded = st.file_uploader(
            "Add images to the pool",
            accept_multiple_files=True,
            type=["jpg", "jpeg", "png", "webp"],
            key=f"pool_upload_{batch_id}",
        )
        if uploaded:
            sig = tuple((f.name, f.size) for f in uploaded)
            if st.session_state.get(f"pool_uploadsig_{batch_id}") != sig:
                st.session_state[f"pool_uploadsig_{batch_id}"] = sig
                for f in uploaded:
                    add_to_image_pool(batch_id, f.name, f.getvalue())
                st.success(f"Added {len(uploaded)} image(s) to the pool.")
                st.rerun()

        pool_images = list_image_pool(batch_id)
        if not pool_images:
            st.info("Pool is empty. Upload images above to populate it.")
            return

        if not products:
            st.info(
                f"{len(pool_images)} image(s) in the pool. Create some "
                "product cards before assigning."
            )

        # Render each pool image as a row: thumbnail + attach controls.
        for img in pool_images:
            with st.container():
                cols = st.columns([1, 2, 1, 1])
                cols[0].image(str(img), caption=img.name, width=140)

                if not products:
                    cols[1].caption("(no products yet)")
                    continue

                product_options = {
                    f"{p.id} · {p.product_name or '(unnamed)'}": p.id
                    for p in products
                }
                pick_key = f"pool_pick_{img.name}"
                role_key = f"pool_role_{img.name}"
                label = cols[1].selectbox(
                    "Attach to product",
                    list(product_options.keys()),
                    key=pick_key,
                )
                role = cols[2].selectbox(
                    "Role",
                    list(REFERENCE_ROLES),
                    key=role_key,
                )

                if cols[3].button("Attach", key=f"pool_attach_{img.name}"):
                    target_pid = product_options[label]
                    product = next((p for p in products if p.id == target_pid), None)
                    if product is None:
                        st.error("Product no longer exists. Refresh.")
                    else:
                        already_taken = role in product.taken_roles()
                        confirm_key = f"pool_overwrite_{img.name}_{target_pid}_{role}"
                        if already_taken and not st.session_state.get(confirm_key):
                            st.warning(
                                f"Product {target_pid} already has a `{role}` "
                                f"reference. Click Attach again to overwrite."
                            )
                            st.session_state[confirm_key] = True
                        else:
                            attach_pool_image_to_product(batch_id, product, img, role)
                            save_batch_products(batch_id, products)
                            st.session_state.pop(confirm_key, None)
                            # Also clear widget state for the keys
                            st.success(
                                f"Attached {img.name} → product {target_pid} ({role})."
                            )
                            st.rerun()

        # Per-image deletion
        st.divider()
        if st.checkbox("Show pool maintenance", key=f"pool_admin_{batch_id}"):
            for img in pool_images:
                cols = st.columns([3, 1])
                cols[0].text(img.name)
                if cols[1].button("Remove from pool", key=f"pool_rm_{img.name}"):
                    remove_from_image_pool(img)
                    st.rerun()


# ---------------------------------------------------------------------------
# Page entrypoint
# ---------------------------------------------------------------------------


def _batch_summary_metrics(products: list[ProductCard]) -> None:
    total = len(products)
    missing_image = sum(1 for p in products if not p.has_image())
    missing_prompts = sum(1 for p in products if not p.has_prompts())
    ready = sum(1 for p in products if p.is_ready_to_export())
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Total", total)
    c2.metric("Missing image", missing_image)
    c3.metric("Missing prompts", missing_prompts)
    c4.metric("Ready to export", ready)


def _filter_products(
    products: list[ProductCard], filter_choice: str
) -> list[ProductCard]:
    if filter_choice == "Missing images":
        return [p for p in products if not p.has_image()]
    if filter_choice == "Missing prompts":
        return [p for p in products if not p.has_prompts()]
    if filter_choice == "Ready to export":
        return [p for p in products if p.is_ready_to_export()]
    return products  # "All"


def page_ai_intake() -> None:
    st.title("AI Product Intake")

    # --- Batch selector ---
    batches = list_batches()
    current = _current_batch()
    if current and current not in batches:
        current = None
    if not batches:
        st.info("No batches yet. Create your first one to get started.")
        if st.button("Create batch", type="primary"):
            new_id = create_batch()
            st.session_state["current_batch"] = new_id
            st.rerun()
        return

    col_sel, col_new = st.columns([3, 1])
    idx = batches.index(current) if current in batches else 0
    selected = col_sel.selectbox("Current batch", batches, index=idx)
    st.session_state["current_batch"] = selected
    if col_new.button("New batch"):
        new_id = create_batch()
        st.session_state["current_batch"] = new_id
        st.rerun()

    batch_id = selected
    st.caption(f"Batch dir: `data/batches/{batch_id}`")

    # --- AI provider config ---
    _ai_provider_panel()
    # OpenRouter blank-model UI warning.
    if os.environ.get("AI_PROVIDER", "").lower() == "openrouter":
        if not (os.environ.get("OPENROUTER_MODEL") or "").strip():
            st.warning(
                "No OpenRouter model set. OpenRouter account/default routing "
                "will be used (`openrouter/auto`)."
            )

    products = load_batch_products(batch_id)

    # --- Bulk add (Part 1) ---
    _bulk_add_section(batch_id, products)
    # `_bulk_add_section` may have rerun. Reload from disk after section.
    products = load_batch_products(batch_id)

    # --- Image pool (Part 2) ---
    _image_pool_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- Global clipboard intake (Part 3 fallback) ---
    _global_clipboard_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- Batch summary ---
    st.subheader(f"Products in batch ({len(products)})")
    _batch_summary_metrics(products)

    cf_col, add_col = st.columns([3, 1])
    filter_choice = cf_col.selectbox(
        "Filter",
        ["All", "Missing images", "Missing prompts", "Ready to export"],
        key="product_filter",
    )
    if add_col.button("+ Add product", type="primary"):
        new_p = ProductCard(id=new_product_id(), status=PRODUCT_STATUS_DRAFT)
        products.append(new_p)
        save_batch_products(batch_id, products)
        st.session_state["expanded_product_id"] = new_p.id
        st.rerun()

    shown = _filter_products(products, filter_choice)
    if not products:
        st.info('Empty batch. Use "Bulk add" or "+ Add product".')
    elif not shown:
        st.info(f"No products match the filter: {filter_choice}.")
    else:
        expand_for = st.session_state.get("expanded_product_id")
        for p in shown:
            chips = " ".join([
                _status_chip("url", bool(p.tiktok_url)),
                _status_chip("img", p.has_image()),
                _status_chip("ip",  bool(p.image_prompt)),
                _status_chip("vp",  bool(p.video_prompt)),
                _status_chip("ready", p.is_ready_to_export()),
            ])
            with st.expander(
                f"{p.id} · {p.product_name or '(unnamed)'} · {chips}",
                expanded=(p.id == expand_for),
            ):
                _render_product_card(batch_id, p, products)
        if expand_for is not None:
            st.session_state["expanded_product_id"] = None

    # --- Summary table ---
    if products:
        st.subheader("Summary table")
        st.dataframe(
            [
                {
                    "id": p.id,
                    "name": p.product_name or "(unnamed)",
                    "source": detect_url_source(p.tiktok_url),
                    "refs": len(p.reference_images),
                    "image_prompt": bool(p.image_prompt),
                    "video_prompt": bool(p.video_prompt),
                    "ready": p.is_ready_to_export(),
                    "status": p.status,
                }
                for p in products
            ],
            use_container_width=True,
        )

    # --- Export / Sync ---
    st.divider()
    st.subheader("Export / Sync Manifest")
    ready_count = sum(1 for p in products if p.is_ready_to_export())
    not_ready = len(products) - ready_count
    if not_ready:
        st.caption(
            f"{ready_count} product(s) ready. {not_ready} will be SKIPPED "
            "(missing image or prompts)."
        )
    sync_to_inputs = st.checkbox(
        "Also sync to inputs/prompt_manifest.md + inputs/reference_images/",
        value=True,
        help=(
            "Copies the generated manifest and its reference images into "
            "the locations the existing CLI flow expects. Leave on unless "
            "you only want to write to the batch folder."
        ),
    )
    if st.button("Export manifest"):
        summary = export_manifest(batch_id, SETTINGS, sync_to_inputs=sync_to_inputs)
        st.success(
            f"Exported {summary['exported_count']} product(s). "
            f"Skipped {summary['skipped_count']} (no prompts or no reference image)."
        )
        st.json(summary)


def _global_clipboard_section(
    batch_id: str, products: list[ProductCard]
) -> None:
    if not HAS_PASTE_BUTTON:
        with st.expander("Clipboard image intake", expanded=False):
            st.info(
                "The `streamlit-paste-button` package isn't installed. "
                "Rebuild the image (`docker compose build`) to enable "
                "the paste button, or keep using per-card uploads."
            )
        return

    with st.expander("Clipboard image intake (paste once → assign)", expanded=False):
        st.caption(
            "If your browser allows clipboard access, click the button below, "
            "then paste (or grant the prompt). The image is staged in this "
            "section until you assign it to a product."
        )
        paste_result = _paste_image_button(
            "📋 Paste image from clipboard",
            key="global_paste_btn",
            errors="ignore",
        )
        if paste_result and getattr(paste_result, "image_data", None) is not None:
            st.session_state["global_pasted_image"] = paste_result.image_data

        staged = st.session_state.get("global_pasted_image")
        if staged is None:
            return

        cols = st.columns([1, 2, 1])
        cols[0].image(staged, caption="Staged paste", width=180)
        if not products:
            cols[1].info("Create a product card first, then return here to attach.")
            return

        product_options = {
            f"{p.id} · {p.product_name or '(unnamed)'}": p.id
            for p in products
        }
        target_label = cols[1].selectbox(
            "Attach to product",
            list(product_options.keys()),
            key="global_paste_target",
        )
        if cols[2].button("Attach", key="global_paste_attach"):
            pid = product_options[target_label]
            target = next((p for p in products if p.id == pid), None)
            if target is None:
                st.error("Product not found.")
                return
            ok, msg = _save_pasted_image_for_product(
                batch_id, target, products, staged
            )
            if ok:
                st.success(msg)
                st.session_state.pop("global_pasted_image", None)
                st.rerun()
            else:
                st.error(msg)


# ===========================================================================
# Guided workflow pages (Phase 3 UI cleanup)
#
# Each guided page focuses on ONE step of the daily flow:
#   1. Dashboard         — overview + next-action card
#   2. Product Intake    — bulk add + manual add
#   3. Images            — image pool + clipboard + per-product status
#   4. AI Prompts        — batch AI generation + per-product edit
#   5. Export Manifest   — readiness table + sync to inputs/
#   6. Flow Batch Run    — Prepare+Generate, Sync Favorites, Generate Videos
#   7. Logs              — today's log + recent error screenshots
# The old all-in-one and per-tool pages stay reachable under [Advanced]
# entries so existing muscle memory and debugging tools keep working.
# ===========================================================================


# --- Goto-page glue --------------------------------------------------------
# Dashboard's "Next action" buttons request a page switch by writing the
# target label into st.session_state["force_page"]. We pop it BEFORE the
# sidebar radio renders so the radio picks it up as its current value.

def _force_navigate(target: str) -> None:
    st.session_state["force_page"] = target
    st.rerun()


# --- Next-action recommender ------------------------------------------------


def _recommend_next_action(batch_id: str | None) -> tuple[str, str, str, str | None]:
    """Decide what the user should do next based on batch + CSV state.

    Returns (severity, label, message, target_page) where severity is
    "info" / "success" / "warning" and target_page (if non-None) is the
    PAGES key the dashboard's button should navigate to.
    """
    if not batch_id:
        return (
            "warning",
            "Create a batch",
            "Start by creating a new batch on the Product Intake page.",
            "2. Product Intake",
        )

    try:
        products = load_batch_products(batch_id)
    except Exception:  # noqa: BLE001
        products = []

    if not products:
        return (
            "info",
            "Add products",
            "The current batch is empty. Paste a JSON list of products on "
            "the Product Intake page, or add one manually.",
            "2. Product Intake",
        )

    missing_images = [p for p in products if not p.has_image()]
    if missing_images:
        return (
            "info",
            f"Add reference images ({len(missing_images)} missing)",
            "Use the Image Pool or per-card paste on the Images page.",
            "3. Images",
        )

    missing_prompts = [p for p in products if not p.has_prompts()]
    if missing_prompts:
        return (
            "info",
            f"Generate AI prompts ({len(missing_prompts)} missing)",
            "Run the batch generator on the AI Prompts page.",
            "4. AI Prompts",
        )

    ready = [p for p in products if p.is_ready_to_export()]
    # Has anyone been exported yet? (status flipped to EXPORTED by export_manifest)
    any_exported = any(p.status == PRODUCT_STATUS_EXPORTED for p in products)
    if ready and not any_exported:
        return (
            "info",
            f"Export + sync manifest ({len(ready)} ready)",
            "Write prompt_manifest.md and sync into inputs/ on the "
            "Export Manifest page.",
            "5. Export Manifest",
        )

    # Look at runtime CSV state.
    total_csv, status_counts = csv_status_counts()
    if total_csv == 0:
        # Exported flag set but CSV empty — manifest hasn't been --load-manifest'd yet.
        if any_exported:
            return (
                "info",
                "Load + generate images",
                "Manifest exported but products.csv is empty. Run "
                '"Prepare + Generate Images" on the Flow Batch Run page.',
                "6. Flow Batch Run",
            )
    else:
        if status_counts.get(STATUS_PENDING, 0) > 0:
            return (
                "info",
                f"Generate images ({status_counts[STATUS_PENDING]} pending)",
                'Run "Prepare + Generate Images" on the Flow Batch Run page.',
                "6. Flow Batch Run",
            )
        if status_counts.get(STATUS_IMAGE_SUBMITTED, 0) > 0:
            return (
                "info",
                f"Heart favorites + sync ({status_counts[STATUS_IMAGE_SUBMITTED]} awaiting)",
                "Heart the good images in your real Chrome Flow window, then "
                "click Sync Favorites on the Flow Batch Run page.",
                "6. Flow Batch Run",
            )
        if status_counts.get(STATUS_IMAGE_APPROVED, 0) > 0:
            return (
                "info",
                f"Generate videos ({status_counts[STATUS_IMAGE_APPROVED]} approved)",
                'Click "Generate Videos for Approved Products" on the Flow '
                "Batch Run page.",
                "6. Flow Batch Run",
            )
        if status_counts.get(STATUS_VIDEO_SUBMITTED, 0) > 0:
            return (
                "success",
                "Batch (mostly) done",
                f"{status_counts[STATUS_VIDEO_SUBMITTED]} video(s) submitted. "
                "Watch them finish in Flow Labs.",
                None,
            )

    return (
        "success",
        "Nothing pending",
        "All products in this batch have moved past the standard pipeline.",
        None,
    )


def _render_next_action_card(batch_id: str | None) -> None:
    severity, label, message, target = _recommend_next_action(batch_id)
    box = {"info": st.info, "warning": st.warning, "success": st.success}[severity]
    box(f"**Next action — {label}**\n\n{message}")
    if target:
        if st.button(f"Go to {target}", key=f"goto_{target}", type="primary"):
            _force_navigate(target)


# --- Compact product list --------------------------------------------------


def _compact_product_row(
    product: ProductCard,
    *,
    show_thumb: bool = True,
    edit_target_page: str | None = None,
) -> None:
    cols = st.columns([1, 4, 3, 1])
    if show_thumb and product.reference_images:
        ref = product.reference_images[0]
        full = (Path(ref) if Path(ref).is_absolute() else (ROOT / ref))
        if full.exists():
            cols[0].image(str(full), width=72)
        else:
            cols[0].caption("(missing)")
    else:
        cols[0].caption("(no img)")

    name = product.product_name or "(unnamed)"
    cols[1].markdown(f"**{name}**")
    if product.tiktok_url:
        cols[1].caption(product.tiktok_url[:90])

    cols[2].markdown(
        " ".join([
            _status_chip("url", bool(product.tiktok_url)),
            _status_chip("img", product.has_image()),
            _status_chip("ip", bool(product.image_prompt)),
            _status_chip("vp", bool(product.video_prompt)),
            _status_chip("ready", product.is_ready_to_export()),
        ])
    )

    if edit_target_page:
        if cols[3].button("Edit", key=f"row_edit_{product.id}_{edit_target_page}"):
            st.session_state["expanded_product_id"] = product.id
            _force_navigate(edit_target_page)


# --- Batch AI generator ----------------------------------------------------


def _batch_generate_ai_prompts(
    batch_id: str,
    products: list[ProductCard],
    *,
    overwrite: bool,
    include_without_images: bool,
) -> None:
    """Iterate eligible products, call the provider for each, save after every one."""
    provider_name = (os.environ.get("AI_PROVIDER") or "manual").lower()
    if provider_name == "manual":
        st.warning(
            "Manual provider selected. Switch AI_PROVIDER (or pick from the "
            "AI provider panel on the AI Prompts page) before running the batch."
        )
        return
    try:
        provider = get_provider(provider_name)
    except Exception as exc:  # noqa: BLE001
        st.error(f"Could not load provider {provider_name}: {exc}")
        return
    ok, msg = provider.is_configured()
    if not ok:
        st.error(f"AI provider not configured: {msg}")
        return

    candidates: list[ProductCard] = []
    for p in products:
        if not (p.product_name.strip() or p.product_description.strip()):
            continue
        if not p.has_image() and not include_without_images:
            continue
        if p.has_prompts() and not overwrite:
            continue
        candidates.append(p)

    if not candidates:
        st.info(
            "No products need generation. "
            'Tick "Overwrite existing prompts" or "Include products without '
            'images" to widen the set.'
        )
        return

    st.write(f"Generating prompts for **{len(candidates)}** product(s) via **{provider_name}**.")
    progress = st.progress(0.0)
    status_box = st.empty()
    log_lines: list[str] = []
    log_box = st.empty()
    success = failed = 0

    for i, p in enumerate(candidates, start=1):
        label = p.product_name or p.original_title or "(unnamed)"
        status_box.text(f"[{i}/{len(candidates)}] {label}")
        product_dict = {
            "product_name":         p.product_name,
            "tiktok_url":           p.tiktok_url,
            "product_description":  p.product_description,
            "notes":                p.notes,
            "reference_filenames":  [Path(r).name for r in p.reference_images],
            "category_hint":        p.category,
            "store_hint":           p.store_environment,
            "placement_hint":       p.placement_type,
        }
        try:
            output = provider.generate_product_prompts(product_dict)
            ok_v, _problems = validate_ai_output(output)
            for field_name in (
                "image_prompt", "video_prompt", "hook", "caption",
                "category", "store_environment", "placement_type",
            ):
                if output.get(field_name):
                    setattr(p, field_name, str(output[field_name]).strip())
            raw_warnings = output.get("warnings") or []
            p.warnings = (
                [str(w) for w in raw_warnings]
                if isinstance(raw_warnings, list)
                else [str(raw_warnings)]
            )
            if ok_v and p.has_prompts():
                p.status = PRODUCT_STATUS_READY
                success += 1
                log_lines.append(f"  OK   {label}")
            else:
                # Partial success — keep status as-is, record as failed.
                failed += 1
                log_lines.append(f"  WARN {label} (validation issues)")
            save_batch_products(batch_id, products)
        except Exception as exc:  # noqa: BLE001
            failed += 1
            log_lines.append(f"  FAIL {label}: {exc}")
        progress.progress(i / len(candidates))
        log_box.code("\n".join(log_lines[-30:]), language="text")

    status_box.empty()
    progress.empty()
    skipped = len(products) - len(candidates)
    summary = (
        f"Done. Generated **{success}** / failed **{failed}** / "
        f"skipped **{skipped}** (didn't meet eligibility)."
    )
    if failed:
        st.warning(summary)
    else:
        st.success(summary)


# --- Pipeline runner -------------------------------------------------------


def _run_pipeline_steps(steps: list[tuple]) -> bool:
    """Run a sequence of CLI steps in order.

    Each step is `(label, args)` or `(label, args, abort_on_fail)`. When
    `abort_on_fail` is False (default True), a non-zero exit logs a
    warning and the pipeline continues — useful for the validate step,
    which exits 1 even when the manifest has only minor warnings that
    don't actually break downstream `--load-manifest` / `--generate-images`.

    Returns True if every fatal-on-fail step exited 0.
    """
    for step in steps:
        if len(step) == 2:
            label, args = step
            abort_on_fail = True
        else:
            label, args, abort_on_fail = step
        st.markdown(f"### {label}")
        rc, _ = run_cli(args)
        if rc != 0:
            if abort_on_fail:
                st.error(f"`{label}` failed (exit {rc}). Pipeline aborted.")
                return False
            st.warning(
                f"`{label}` exited {rc}. Treating as a warning and "
                "continuing — `--validate-manifest` returns non-zero for "
                "any per-section issue, not just fatal ones."
            )
    return True


# --- Page 1: Dashboard (replaces previous page_dashboard) ------------------


def page_guided_dashboard() -> None:
    st.title("Dashboard")

    batches = list_batches()
    if not batches:
        st.info("No batches yet.")
        if st.button("Create first batch", type="primary"):
            new_id = create_batch()
            st.session_state["current_batch"] = new_id
            _force_navigate("2. Product Intake")
        return

    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current

    st.markdown(f"**Current batch:** `{current}`")
    _render_next_action_card(current)

    st.divider()
    st.subheader("Connectivity")
    ok, info = chrome_reachable()
    if ok:
        st.success(f"Chrome reachable via {chrome_cdp_url()}")
    else:
        st.error(f"Cannot reach Chrome CDP at {chrome_cdp_url()}.")
        st.caption(info[:300])

    st.subheader(f"Batch counts (`{current}`)")
    products = load_batch_products(current)
    total = len(products)
    with_images = sum(1 for p in products if p.has_image())
    with_prompts = sum(1 for p in products if p.has_prompts())
    ready = sum(1 for p in products if p.is_ready_to_export())

    csv_total, csv_counts = csv_status_counts()

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Products", total)
    c2.metric("With images", with_images)
    c3.metric("With prompts", with_prompts)
    c4.metric("Ready to export", ready)

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("CSV rows", csv_total)
    c6.metric("image_submitted", csv_counts.get(STATUS_IMAGE_SUBMITTED, 0))
    c7.metric("image_approved",  csv_counts.get(STATUS_IMAGE_APPROVED, 0))
    c8.metric("video_submitted", csv_counts.get(STATUS_VIDEO_SUBMITTED, 0))


# --- Page 2: Product Intake -----------------------------------------------


def page_product_intake() -> None:
    st.title("Product Intake")

    # Batch selector at top of every guided page.
    batches = list_batches()
    current = _current_batch()
    if not batches:
        st.info("No batches yet.")
        if st.button("Create batch", type="primary"):
            new_id = create_batch()
            st.session_state["current_batch"] = new_id
            st.rerun()
        return
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current

    cols = st.columns([3, 1])
    selected = cols[0].selectbox("Current batch", batches, index=batches.index(current))
    st.session_state["current_batch"] = selected
    if cols[1].button("New batch"):
        new_id = create_batch()
        st.session_state["current_batch"] = new_id
        st.rerun()
    batch_id = selected

    _render_next_action_card(batch_id)

    products = load_batch_products(batch_id)

    # --- PRIMARY: Kalodata Excel import ---
    st.subheader("Import from Kalodata Excel (preferred)")
    _kalodata_import_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- Bulk JSON add (kept) ---
    st.subheader("Bulk add via JSON")
    _bulk_add_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- SECONDARY: manual single add ---
    with st.expander("Add one product manually", expanded=False):
        st.caption(
            "Adds a blank card you can fill in from scratch. "
            "Use Bulk Add when you have a list of TikTok URLs."
        )
        if st.button("+ Add blank product card"):
            new_p = ProductCard(id=new_product_id(), status=PRODUCT_STATUS_DRAFT)
            products.append(new_p)
            save_batch_products(batch_id, products)
            st.session_state["expanded_product_id"] = new_p.id
            st.success(f"Added blank product {new_p.id}. Edit it on the AI Prompts page.")
            st.rerun()

    # --- Products list (compact) ---
    st.subheader(f"Products in batch ({len(products)})")
    if not products:
        st.info("Empty batch — use Bulk Add above.")
    else:
        for p in products:
            _compact_product_row(p, edit_target_page="4. AI Prompts")
            st.divider()


# --- Page 3: Images -------------------------------------------------------


def page_images() -> None:
    st.title("Images")

    batches = list_batches()
    if not batches:
        st.info("No batches yet — create one on the Product Intake page.")
        return
    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current
    batch_id = current
    st.markdown(f"**Current batch:** `{batch_id}`")

    products = load_batch_products(batch_id)
    _render_next_action_card(batch_id)

    # --- PRIMARY: image pool ---
    _image_pool_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- Global clipboard intake ---
    _global_clipboard_section(batch_id, products)
    products = load_batch_products(batch_id)

    # --- Per-product status + filter ---
    st.subheader("Per-product image status")
    if not products:
        st.info("No products yet.")
        return

    filter_choice = st.selectbox(
        "Filter",
        ["All products", "Missing images", "Has images"],
        key="images_filter",
    )
    if filter_choice == "Missing images":
        shown = [p for p in products if not p.has_image()]
    elif filter_choice == "Has images":
        shown = [p for p in products if p.has_image()]
    else:
        shown = products

    if not shown:
        st.info(f"No products match: {filter_choice}.")
        return

    for p in shown:
        _compact_product_row(p, edit_target_page="3. Images")
        st.divider()

    # --- Per-card editor (paste / upload) ---
    expand_for = st.session_state.get("expanded_product_id")
    if expand_for:
        target = next((p for p in products if p.id == expand_for), None)
        if target is not None:
            st.subheader(f"Edit: {target.product_name or target.id}")
            _render_product_card(batch_id, target, products)


# --- Page 4: AI Prompts ---------------------------------------------------


def page_ai_prompts() -> None:
    st.title("AI Prompts")

    batches = list_batches()
    if not batches:
        st.info("No batches yet.")
        return
    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current
    batch_id = current
    st.markdown(f"**Current batch:** `{batch_id}`")

    products = load_batch_products(batch_id)
    _render_next_action_card(batch_id)

    _ai_provider_panel()
    if (os.environ.get("AI_PROVIDER") or "").lower() == "openrouter":
        if not (os.environ.get("OPENROUTER_MODEL") or "").strip():
            st.warning(
                "No OpenRouter model set. OpenRouter account/default "
                "routing will be used (`openrouter/auto`)."
            )

    # --- BATCH GENERATION ---
    st.subheader("Generate AI prompts for all ready products")
    overwrite = st.checkbox(
        "Overwrite existing prompts",
        value=False,
        key="batch_ai_overwrite",
        help="If unchecked, products that already have an image_prompt + "
             "video_prompt are skipped.",
    )
    include_no_image = st.checkbox(
        "Include products without images",
        value=False,
        key="batch_ai_include_no_image",
        help="If unchecked (default), products with no reference image "
             "are skipped — Flow needs the image to generate anyway.",
    )
    if (os.environ.get("AI_PROVIDER") or "").lower() == "manual":
        st.info("Manual provider selected. Enter prompts manually per card below.")
    else:
        if st.button(
            "Generate AI Prompts for All Ready Products",
            type="primary",
            key="batch_ai_go",
        ):
            _batch_generate_ai_prompts(
                batch_id, products,
                overwrite=overwrite,
                include_without_images=include_no_image,
            )
            st.rerun()

    # --- Per-product compact + expand-on-edit ---
    st.subheader(f"Products in batch ({len(products)})")
    filter_choice = st.selectbox(
        "Filter",
        ["All products", "Missing prompts", "Ready to export"],
        key="ai_filter",
    )
    if filter_choice == "Missing prompts":
        shown = [p for p in products if not p.has_prompts()]
    elif filter_choice == "Ready to export":
        shown = [p for p in products if p.is_ready_to_export()]
    else:
        shown = products

    if not shown:
        st.info(f"No products match: {filter_choice}.")
    else:
        for p in shown:
            _compact_product_row(p, edit_target_page="4. AI Prompts")
            st.divider()

    expand_for = st.session_state.get("expanded_product_id")
    if expand_for:
        target = next((p for p in products if p.id == expand_for), None)
        if target is not None:
            st.subheader(f"Edit: {target.product_name or target.id}")
            _render_product_card(batch_id, target, products)


# --- Page 5: Export Manifest ----------------------------------------------


def page_export_manifest() -> None:
    st.title("Export Manifest")

    batches = list_batches()
    if not batches:
        st.info("No batches yet.")
        return
    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current
    batch_id = current
    st.markdown(f"**Current batch:** `{batch_id}`")

    products = load_batch_products(batch_id)
    _render_next_action_card(batch_id)

    if not products:
        st.info("No products in batch.")
        return

    # Readiness table with skip reasons.
    rows = []
    for p in products:
        reasons = []
        if not p.product_name.strip():
            reasons.append("missing product_name")
        if not p.has_image():
            reasons.append("missing reference image")
        if not p.image_prompt.strip():
            reasons.append("missing image_prompt")
        if not p.video_prompt.strip():
            reasons.append("missing video_prompt")
        rows.append({
            "id":            p.id,
            "name":          p.product_name or "(unnamed)",
            "has_image":     p.has_image(),
            "has_img_pr":    bool(p.image_prompt),
            "has_vid_pr":    bool(p.video_prompt),
            "ready":         p.is_ready_to_export(),
            "skip_reason":   ", ".join(reasons) if reasons else "",
        })
    st.dataframe(rows, use_container_width=True)

    ready_count = sum(1 for p in products if p.is_ready_to_export())
    skipped_count = len(products) - ready_count

    include_incomplete = st.checkbox(
        "Include incomplete products",
        value=False,
        help="Export will normally skip products with missing image / "
             "prompt fields. Tick this only if you've fixed them by hand "
             "and want to push them through anyway.",
    )
    sync_to_inputs = st.checkbox(
        "Sync to inputs/prompt_manifest.md + inputs/reference_images/",
        value=True,
        help="The existing CLI flow reads from inputs/. Leave on for the "
             "normal daily flow.",
    )

    st.caption(
        f"{ready_count} ready, {skipped_count} would be skipped "
        f"(uncheck above to include them anyway)."
    )

    if st.button("Export + Sync Manifest", type="primary"):
        # `include_incomplete` doesn't currently bypass the readiness check
        # inside export_manifest — it requires `image_prompt + video_prompt
        # + primary reference image`. We surface that here so the user
        # knows.
        if include_incomplete and skipped_count > 0:
            st.info(
                "Note: products that lack a reference image or any of the "
                "two prompts can't be exported by the manifest format. "
                "They will still be skipped."
            )
        summary = export_manifest(batch_id, SETTINGS, sync_to_inputs=sync_to_inputs)
        st.success(
            f"Exported {summary['exported_count']} product(s). "
            f"Skipped {summary['skipped_count']}."
        )
        st.json(summary)


# --- Page 6: Flow Batch Run ----------------------------------------------


def page_flow_batch_run() -> None:
    st.title("Flow Batch Run")

    batches = list_batches()
    if not batches:
        st.info("No batches yet.")
        return
    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current
    st.markdown(f"**Current batch:** `{current}`")

    _render_next_action_card(current)

    default_limit = st.session_state.get("batch_limit", 30)
    limit = st.number_input(
        "Limit (rows per batch)",
        min_value=1, max_value=200, value=default_limit, step=1,
        key="flow_run_limit",
    )
    st.session_state["batch_limit"] = limit

    st.divider()
    # --- Step 6: Prepare + Generate Images ---
    st.subheader("Step 6 — Prepare + Generate Images")
    st.caption(
        "Runs Validate Manifest → Load Manifest Fresh → Generate Images "
        "in that order. Stops at the first failure so you can see exactly "
        "where it broke."
    )
    confirm_fresh = st.checkbox(
        "I understand this will run --load-manifest --fresh "
        "(backs up the current CSV).",
        value=False,
        key="confirm_fresh_prepare",
    )
    if st.button(
        f"Prepare + Generate Images (limit {limit})",
        type="primary",
        disabled=not confirm_fresh,
    ):
        _run_pipeline_steps([
            # Validate is informational here — non-fatal. The CLI's
            # exit-1-on-any-warning is too strict for the pipeline, so
            # we surface its output but don't abort on it.
            ("Validate Manifest", ["--validate-manifest", str(MANIFEST_PATH)], False),
            ("Load Manifest Fresh", ["--load-manifest", str(MANIFEST_PATH), "--fresh"], True),
            (f"Generate Images (limit {limit})", ["--generate-images", "--limit", str(limit)], True),
        ])

    with st.expander("Run sub-steps individually", expanded=False):
        c1, c2 = st.columns(2)
        if c1.button("Check Browser", key="frb_check"):
            run_cli(["--check-browser"])
        if c2.button("Validate Manifest", key="frb_validate"):
            run_cli(["--validate-manifest", str(MANIFEST_PATH)])
        c3, c4 = st.columns(2)
        if c3.button("Load Manifest Fresh", key="frb_load",
                     disabled=not st.checkbox("confirm fresh", key="conf_fresh_sub")):
            run_cli(["--load-manifest", str(MANIFEST_PATH), "--fresh"])
        if c4.button(f"Generate Images (limit {limit})", key="frb_gen"):
            run_cli(["--generate-images", "--limit", str(limit)])

    st.divider()
    # --- Step 7: Sync Favorites ---
    st.subheader("Step 7 — Review favorites + sync")
    st.info(
        "Review the generated images in your real Chrome Flow window. "
        "Heart / favorite the ones you want to animate. Then click "
        "Sync Favorites below."
    )
    if st.button("Sync Favorites", type="primary"):
        run_cli(["--sync-favorites"])

    csv_total, csv_counts = csv_status_counts()
    if csv_total > 0:
        st.write(
            f"Submitted: {csv_counts.get(STATUS_IMAGE_SUBMITTED, 0)} · "
            f"Approved: {csv_counts.get(STATUS_IMAGE_APPROVED, 0)} · "
            f"Video submitted: {csv_counts.get(STATUS_VIDEO_SUBMITTED, 0)}"
        )

    st.divider()
    # --- Step 8: Generate Videos ---
    st.subheader("Step 8 — Generate videos for approved products")
    if st.button(
        f"Generate Videos for Approved Products (limit {limit})",
        type="primary",
        key="frb_videos",
    ):
        run_cli(["--generate-videos", "--limit", str(limit)])


# ===========================================================================
# Sidebar wiring
# ===========================================================================


# ===========================================================================
# BOF Batch Builder — the single-page main workflow (Phase 3 UI cleanup #2)
#
# The whole daily flow lives on this page. Technical pipeline terms
# (manifest, CSV, --load-manifest, --validate-manifest) are hidden:
# the user sees business-level actions like "Prepare image batch" and
# "Generate videos for favorited images". Every internal step is still
# accessible from the Advanced expander in the sidebar for debugging.
# ===========================================================================


def _user_next_action(batch_id: str | None) -> tuple[str, str]:
    """User-facing next-action recommendation. Returns (severity, message)."""
    if not batch_id:
        return "info", "Upload a Kalodata export or add products manually."
    try:
        products = load_batch_products(batch_id)
    except Exception:  # noqa: BLE001
        products = []
    if not products:
        return "info", "Upload a Kalodata export or add products manually."

    missing_images = sum(1 for p in products if not p.has_image())
    if missing_images:
        return "info", f"Add missing reference images ({missing_images} products)."

    missing_prompts = sum(1 for p in products if not p.has_prompts())
    if missing_prompts:
        return "info", f"Generate prompts for all products ({missing_prompts} need prompts)."

    ready = [p for p in products if p.is_ready_to_export()]
    any_exported = any(p.status == PRODUCT_STATUS_EXPORTED for p in products)
    if ready and not any_exported:
        return "info", "Prepare image batch."

    csv_total, csv_counts = csv_status_counts()
    if csv_total == 0 and any_exported:
        return "info", "Image batch ready. Generate images in Flow."
    if csv_counts.get(STATUS_PENDING, 0) > 0:
        return "info", "Generate images in Flow."
    if csv_counts.get(STATUS_IMAGE_SUBMITTED, 0) > 0:
        return "info", "Heart/favorite good images in Flow, then sync favorites."
    if csv_counts.get(STATUS_IMAGE_APPROVED, 0) > 0:
        return "info", "Generate videos for favorited images."
    if csv_counts.get(STATUS_VIDEO_SUBMITTED, 0) > 0:
        return "success", "Videos are submitted — they're rendering in Flow."
    return "success", "All products processed."


def _render_unmatched_favorites_section(batch_id: str) -> None:
    """Step 6 — show favorited Flow tiles that didn't auto-bind to a product.

    Lets the user pick a product from the current batch and bind the
    favorite to it. Binding promotes the media_id to the front of the
    row's `flow_media_id` list and flips the status to image_approved.
    """
    items = load_unmatched_favorites()
    if not items:
        return

    st.markdown(f"#### Favorited images that need review ({len(items)})")
    st.caption(
        "These favorited tiles weren't matched to a product in your "
        "batch — most often because you regenerated a variant in Flow "
        "by hand. Pick a product to attach each one to. Video "
        "generation can also animate them directly without binding."
    )

    products = load_batch_products(batch_id)
    if not products:
        st.info(
            "No products in this batch to bind to. Switch to the batch "
            "that owns these favorites, or import the products first."
        )
        return

    product_options = {
        f"{p.id} — {p.product_name or '(unnamed)'}": p.id for p in products
    }

    for it in items:
        with st.container():
            cols = st.columns([1, 2, 2, 1])
            # Render the thumbnail as an inline <img> instead of using
            # st.image, because st.image makes the Streamlit container
            # fetch the URL — and the container isn't authenticated with
            # labs.google. The browser, however, has the user's session
            # cookie, so an inline <img> loads fine.
            if it.flow_image_src:
                img_href = it.flow_image_src
                if img_href.startswith("/"):
                    img_href = "https://labs.google" + img_href
                edit_href = it.tile_href
                if edit_href and edit_href.startswith("/"):
                    edit_href = "https://labs.google" + edit_href
                html = (
                    f'<img src="{img_href}" '
                    f'style="width:120px;height:auto;border-radius:6px;'
                    f'background:#222;object-fit:cover" '
                    f'alt="favorited tile" '
                    f'referrerpolicy="no-referrer-when-downgrade"/>'
                )
                if edit_href:
                    html = f'<a href="{edit_href}" target="_blank">{html}</a>'
                cols[0].markdown(html, unsafe_allow_html=True)
            else:
                cols[0].caption("(no thumb)")

            cols[1].markdown(f"**media_id**\n\n`{it.media_id[:18]}…`")
            cols[1].caption(
                f"tile_id: {it.flow_tile_id[:16] or '—'}  ·  "
                f"edit_id: {it.edit_id[:8] or '—'}"
            )
            if it.tile_href:
                edit_href = it.tile_href
                if edit_href.startswith("/"):
                    edit_href = "https://labs.google" + edit_href
                cols[1].markdown(f"[Open in Flow]({edit_href})")

            picked = cols[2].selectbox(
                "Bind to product",
                ["(select)"] + list(product_options.keys()),
                key=f"unmatched_pick_{it.media_id}",
            )

            if cols[3].button("Bind", key=f"unmatched_bind_{it.media_id}"):
                if picked == "(select)":
                    st.warning("Pick a product first.")
                else:
                    target_pid = product_options[picked]
                    ok, msg = bind_unmatched_favorite_to_product(
                        SETTINGS,
                        it.media_id,
                        target_pid,
                        logger=__import__("logging").getLogger("ui"),
                    )
                    if ok:
                        st.success(msg)
                        st.rerun()
                    else:
                        st.error(msg)

            if cols[3].button("Dismiss", key=f"unmatched_dismiss_{it.media_id}"):
                remove_unmatched_favorite(it.media_id)
                st.rerun()
        st.divider()


def _prepare_image_batch(batch_id: str) -> None:
    """Combined: export manifest -> validate -> load fresh.

    Hides three CLI commands behind one user-facing action. Validate is
    non-fatal (its exit-1 fires on per-row warnings); the load-fresh
    step is the only mandatory one because it's what rebuilds the
    runtime CSV the rest of the pipeline reads.
    """
    st.markdown("### Step A — Export from product cards")
    summary = export_manifest(batch_id, SETTINGS, sync_to_inputs=True)
    st.write(
        f"Exported **{summary['exported_count']}** product(s) to the "
        f"image batch. Skipped {summary['skipped_count']} "
        "(missing image or prompts)."
    )
    if summary["exported_count"] == 0:
        st.error(
            "Nothing to prepare — every product is missing an image or "
            "prompts. Fix incomplete products and try again."
        )
        return

    st.markdown("### Step B — Quick validate (warnings only)")
    rc, _ = run_cli(["--validate-manifest", str(MANIFEST_PATH)])
    if rc != 0:
        st.warning("Some products have warnings (non-fatal). Continuing.")

    st.markdown("### Step C — Initialize batch state")
    rc, _ = run_cli(["--load-manifest", str(MANIFEST_PATH), "--fresh"])
    if rc != 0:
        st.error(
            "Could not initialize the batch state. See the output above "
            "for details. (CLI: `--load-manifest --fresh`)"
        )
        return
    st.success("Image batch is ready. Move to **Step 4 — Generate Images in Flow**.")


def _render_blanket_video_prompt_panel() -> None:
    """Show the universal video prompt + an editable textarea.

    Edits save to data/settings.local.json and apply on the next run.
    """
    from src.config import DEFAULT_BLANKET_VIDEO_PROMPT

    use_blanket = (
        os.environ.get("USE_BLANKET_VIDEO_PROMPT", "true") or "true"
    ).strip().lower() not in {"0", "false", "no", "off"}

    if not use_blanket:
        st.info(
            "USE_BLANKET_VIDEO_PROMPT is disabled. Videos will use each "
            "product's own video_prompt. Set it back to true in **Setup** "
            "to avoid mismatches when images are regenerated or rebound."
        )
        return

    with st.container(border=True):
        st.markdown(
            "**Using blanket video prompt for all products.**  \n"
            "Every product is animated with the same universal prompt so "
            "that regenerated or rebound images can't drift away from a "
            "product-specific video prompt that was authored against an "
            "earlier image."
        )
        current = (
            os.environ.get("BLANKET_VIDEO_PROMPT")
            or DEFAULT_BLANKET_VIDEO_PROMPT
        )
        new_value = st.text_area(
            "Blanket video prompt",
            value=current,
            height=140,
            key="blanket_video_prompt_textarea",
            help=(
                "Used for every video. Edit here and click Save to "
                "persist to data/settings.local.json."
            ),
        )
        cs1, cs2 = st.columns([1, 4])
        if cs1.button("Save", key="save_blanket_prompt"):
            new_value = new_value.strip()
            if not new_value:
                st.error("Blanket prompt can't be empty.")
            else:
                try:
                    saved = load_user_settings()
                    saved.blanket_video_prompt = new_value
                    # Ensure the master switch stays on; the textarea is
                    # only meaningful when blanket mode is enabled.
                    saved.use_blanket_video_prompt = "true"
                    save_user_settings(saved)
                    apply_user_settings_to_env(saved, load_user_secrets())
                    cs2.success("Saved. Active on the next video run.")
                except Exception as exc:  # noqa: BLE001
                    cs2.error(f"Could not save: {exc}")


def page_bof_batch_builder() -> None:
    st.title("BOF Batch Builder")
    st.caption(
        f"End-to-end batch authoring on one page. "
        f"**Prompt mode:** {_prompt_mode_label()} "
        f"(change in **Setup**)."
    )

    if not _ai_provider_is_configured():
        with st.container(border=True):
            st.warning(
                "⚠️  No AI provider is configured. The **AI Prompts** step "
                "below will fail. Open **Setup** in the sidebar to enter "
                "your API key (it's saved locally, never to git)."
            )
            if st.button("Go to Setup", key="bof_goto_setup"):
                st.session_state["force_page"] = "Setup"
                st.rerun()

    # --- Batch picker (kept compact at the top) ---
    batches = list_batches()
    if not batches:
        st.info("No batches yet. Click below to start a new one.")
        if st.button("Create first batch", type="primary"):
            new_id = create_batch()
            st.session_state["current_batch"] = new_id
            st.rerun()
        return

    current = _current_batch()
    if current not in batches:
        current = batches[0]
        st.session_state["current_batch"] = current

    bc1, bc2 = st.columns([4, 1])
    selected = bc1.selectbox(
        "Current batch",
        batches,
        index=batches.index(current),
        key="bof_batch_select",
    )
    st.session_state["current_batch"] = selected
    if bc2.button("New batch"):
        new_id = create_batch()
        st.session_state["current_batch"] = new_id
        st.rerun()
    batch_id = selected

    # --- Next action banner (always at top) ---
    severity, msg = _user_next_action(batch_id)
    box = {"info": st.info, "warning": st.warning, "success": st.success}.get(
        severity, st.info
    )
    box(f"**Next action:** {msg}")

    # --- Automation mode (Phase 4) — overrides the env-loaded default
    #     for THIS Streamlit process. Affects every subprocess we spawn
    #     from now on since they inherit env via os.environ.copy(). ---
    with st.expander("Automation speed", expanded=False):
        # "safe" was retired -- the per-mode sleep deltas didn't change
        # reliability for any of the failure modes we saw in practice
        # (the real culprit is selector_timeout_ms, which is the same
        # value across modes). Legacy AUTOMATION_MODE=safe env values
        # are coerced to balanced in src/config.py:load_settings.
        modes = ["fast", "balanced"]
        current_mode = (os.environ.get("AUTOMATION_MODE") or "fast").lower()
        if current_mode == "safe":
            current_mode = "balanced"
        if current_mode not in modes:
            current_mode = "fast"
        new_mode = st.radio(
            "Mode",
            modes,
            index=modes.index(current_mode),
            horizontal=True,
            help=(
                "fast = default. Shortest waits + fewer retries.  "
                "balanced = a bit slower; use only if you're seeing UI "
                "race conditions on a sluggish machine."
            ),
            key="bof_automation_mode",
        )
        if new_mode != current_mode:
            os.environ["AUTOMATION_MODE"] = new_mode
            # Drop the cached settings so subsequent reads pick up the new mode.
            for k in list(os.environ):
                if k.startswith(("VIDEO_", "IMAGE_BETWEEN", "IMAGE_UI")):
                    # Clear per-knob overrides so the mode defaults apply.
                    if k in (
                        "VIDEO_TILE_SETTLE_MS", "VIDEO_AFTER_HOVER_MS",
                        "VIDEO_AFTER_MENU_CLICK_MS", "VIDEO_BETWEEN_PRODUCTS_MS",
                        "VIDEO_RETRY_COUNT",
                        "IMAGE_BETWEEN_PRODUCTS_MS", "IMAGE_UI_SETTLE_MS",
                    ):
                        if not (os.environ.get(k) or "").strip():
                            continue
            st.success(f"Mode set to **{new_mode}**. Re-running…")
            st.rerun()
        debug_shots = st.checkbox(
            "Save debug screenshots on every step (slower)",
            value=(os.environ.get("DEBUG_SCREENSHOTS", "false").lower() == "true"),
            key="bof_debug_shots",
        )
        os.environ["DEBUG_SCREENSHOTS"] = "true" if debug_shots else "false"

    products = load_batch_products(batch_id)

    # =====================================================================
    # STEP 1 — Upload Products
    # =====================================================================
    st.divider()
    st.header("1. Upload products")

    intake_tab1, intake_tab2 = st.tabs(
        ["Kalodata Export (xlsx)", "URLs + Reference Images"]
    )
    with intake_tab1:
        _kalodata_import_section(batch_id, products)
    with intake_tab2:
        _bulk_add_section(batch_id, products)
        with st.expander("Add one product manually", expanded=False):
            if st.button("+ Add blank product card", key="bof_add_blank"):
                p = ProductCard(id=new_product_id(), status=PRODUCT_STATUS_DRAFT)
                products.append(p)
                save_batch_products(batch_id, products)
                st.session_state["expanded_product_id"] = p.id
                st.rerun()
        _image_pool_section(batch_id, products)
        _global_clipboard_section(batch_id, products)

    # Refresh state after any of the sub-sections might have rerun
    products = load_batch_products(batch_id)

    # =====================================================================
    # STEP 2 — Verify product list
    # =====================================================================
    st.divider()
    st.header("2. Verify product list")
    if not products:
        st.info("No products yet. Use Step 1 to add some.")
    else:
        # Friendly readiness rollup
        rows = []
        for p in products:
            if not p.has_image():
                status = "Missing image"
            elif not p.product_name or not p.product_description.strip():
                status = "Missing product info"
            elif not p.has_prompts():
                status = "Ready for prompts"
            elif p.is_ready_to_export():
                status = "Ready"
            else:
                status = "Incomplete"
            rows.append({
                "id": p.id,
                "product": p.product_name or "(unnamed)",
                "source": detect_url_source(p.tiktok_url) if p.tiktok_url else "Manual",
                "has_image": p.has_image(),
                "has_info": bool(p.product_description.strip() or p.product_name.strip()),
                "status": status,
            })
        st.dataframe(rows, use_container_width=True)

        with st.expander("Edit a product", expanded=False):
            options = {f"{p.id} — {p.product_name or '(unnamed)'}": p.id for p in products}
            picked = st.selectbox(
                "Pick a product to edit",
                ["(none)"] + list(options.keys()),
                key="bof_pick_edit",
            )
            if picked != "(none)":
                target_id = options[picked]
                target = next((p for p in products if p.id == target_id), None)
                if target is not None:
                    _render_product_card(batch_id, target, products)

    # =====================================================================
    # STEP 3 — Generate prompts + prepare image batch
    # =====================================================================
    st.divider()
    st.header("3. Generate prompts + prepare image batch")
    _ai_provider_panel()
    if (os.environ.get("AI_PROVIDER") or "").lower() == "openrouter":
        if not (os.environ.get("OPENROUTER_MODEL") or "").strip():
            st.warning(
                "No OpenRouter model set. OpenRouter's auto-router will be "
                "used (`openrouter/auto`)."
            )

    if (os.environ.get("AI_PROVIDER") or "").lower() == "manual":
        st.info("Manual provider — enter prompts per card in Step 2.")
    else:
        overwrite = st.checkbox(
            "Overwrite existing prompts",
            value=False,
            key="bof_overwrite",
        )
        include_no_image = st.checkbox(
            "Include products without images",
            value=False,
            key="bof_include_noimg",
        )
        if st.button(
            "Generate Prompts for All Products",
            type="primary",
            key="bof_gen_prompts",
        ):
            _batch_generate_ai_prompts(
                batch_id, products,
                overwrite=overwrite,
                include_without_images=include_no_image,
            )
            st.rerun()

    st.markdown("---")
    confirm_prepare = st.checkbox(
        "I'm ready to prepare the image batch "
        "(this rebuilds the internal batch state — old in-flight rows "
        "are backed up).",
        value=False,
        key="bof_confirm_prepare",
    )
    if st.button(
        "Prepare Image Batch",
        type="primary",
        disabled=not confirm_prepare,
        key="bof_prepare",
    ):
        _prepare_image_batch(batch_id)

    # =====================================================================
    # STEP 4 — Run image generation
    # =====================================================================
    st.divider()
    st.header("4. Generate images in Flow")
    img_limit = st.number_input(
        "How many to generate this run",
        min_value=1, max_value=200,
        value=st.session_state.get("batch_limit", 30),
        step=1,
        key="bof_img_limit",
    )
    st.session_state["batch_limit"] = img_limit

    csv_total, csv_counts = csv_status_counts()
    if csv_total == 0:
        st.info("Prepare the image batch (Step 3) before generating.")
    else:
        st.caption(
            f"Current batch state — pending: {csv_counts.get(STATUS_PENDING, 0)} · "
            f"submitted: {csv_counts.get(STATUS_IMAGE_SUBMITTED, 0)}."
        )
    if st.button(
        f"Generate Images in Flow (up to {img_limit})",
        type="primary",
        key="bof_gen_images",
    ):
        run_cli(["--generate-images", "--limit", str(int(img_limit))])

    # =====================================================================
    # STEP 5 — Pick winners
    # =====================================================================
    st.divider()
    st.header("5. Pick winners")
    st.info(
        "Review the generated images in Flow. Heart/favorite the "
        "images you want to turn into videos. Then click **Scan "
        "Favorites** so the app picks them up."
    )
    pw1, pw2 = st.columns([1, 2])
    if pw1.button(
        "Open Flow",
        key="bof_open_flow",
        help="Opens https://labs.google/flow in a new tab.",
    ):
        st.markdown(
            "<script>window.open('https://labs.google/flow','_blank');</script>",
            unsafe_allow_html=True,
        )
        st.caption(
            "If your browser blocked the popup, open "
            "[labs.google/flow](https://labs.google/flow) manually."
        )
    if pw2.button("Scan Favorites", type="primary", key="bof_sync"):
        run_cli(["--sync-favorites"])

    csv_total, csv_counts = csv_status_counts()
    if csv_total > 0:
        ca, cb, cc = st.columns(3)
        ca.metric("Picked (favorited)", csv_counts.get(STATUS_IMAGE_APPROVED, 0))
        cb.metric("Awaiting your review", csv_counts.get(STATUS_IMAGE_SUBMITTED, 0))
        cc.metric("Skipped", csv_counts.get(STATUS_IMAGE_REJECTED, 0))

    # ---- Favorited images that need review (was "Unmatched Favorited Images") ----
    _render_unmatched_favorites_section(batch_id)

    # =====================================================================
    # STEP 6 — Generate videos
    # =====================================================================
    st.divider()
    st.header("6. Generate videos")

    _render_blanket_video_prompt_panel()

    vid_limit = st.number_input(
        "How many videos to submit this run",
        min_value=1, max_value=200,
        value=st.session_state.get("batch_limit", 30),
        step=1,
        key="bof_vid_limit",
    )

    # Counters now describe the favorited-tile world, not CSV rows. We
    # can know how many media_ids have ALREADY been submitted (state
    # file) but the count of "ready to submit" requires an actual Flow
    # scan, which is too expensive to do on every UI render. So we show
    # a deferred "ready" count that the user can refresh on demand.
    from src.video_state import load_submitted_media_ids
    submitted_ids = load_submitted_media_ids()
    ready_count = st.session_state.get("bof_vid_ready_count")
    csv_total, csv_counts = csv_status_counts()
    vid_errors = csv_counts.get("video_error", 0)

    rc_str = "(click 'Scan favorites' to refresh)" if ready_count is None else str(ready_count)
    st.caption(
        f"Favorited images ready for video: **{rc_str}** · "
        f"Videos submitted: **{len(submitted_ids)}** · "
        f"Video errors: **{vid_errors}**"
    )

    include_already = st.checkbox(
        "Include already submitted favorited images",
        value=False,
        key="bof_vid_include_already",
        help=(
            "By default the app skips media_ids it already animated in a "
            "previous run (tracked in data/video_submitted_tiles.json). "
            "Tick this to re-submit them anyway."
        ),
    )

    vc1, vc2 = st.columns([1, 2])
    if vc1.button("Scan favorites", key="bof_vid_scan", help=(
        "Open Flow and count the favorited image tiles you currently "
        "have. Use this if you want a count before you click Generate."
    )):
        run_cli(["--list-unmatched-favorites"])
        st.info(
            "Scan command launched. Look at the log block above for the "
            "favorited tile list. The exact count of *new* tiles to "
            "submit will print at the top of the next Generate run too."
        )

    btn_args = ["--generate-videos", "--limit", str(int(vid_limit))]
    if include_already:
        btn_args.append("--include-already-submitted")
    if vc2.button(
        f"Generate Videos from Favorited Images (up to {vid_limit})",
        type="primary",
        key="bof_gen_videos",
    ):
        run_cli(btn_args)

    # Advanced controls + raw stdout/stderr live on the dedicated
    # **Advanced / Logs** sidebar entry. Everything that used to be
    # collapsed inline here (check-browser, validate-manifest,
    # load-manifest-fresh, sync-favorites legacy, generate-videos from
    # approved rows) is now one click away in that page, with the same
    # button labels.
    st.divider()
    st.caption(
        "Need to inspect raw logs or run a diagnostic? Open "
        "**Advanced / Logs** in the sidebar."
    )


def page_advanced_logs() -> None:
    """Debug surface: latest CLI run + raw CLI buttons + legacy pages.

    Three sections:
      1. Most recent command — exit code, elapsed, full stdout/stderr.
      2. Raw CLI buttons — every internal --flag the BOF page hides.
      3. Legacy pages — direct links to the pre-cleanup multi-page UI
         and the old logs file browser.
    """
    st.title("Advanced / Logs")
    st.caption(
        "Everything below is here for debugging. The normal workflow "
        "lives entirely on **BOF Batch Builder**."
    )

    # --- Latest command run --------------------------------------------
    st.subheader("Most recent command")
    output = st.session_state.get("last_command_output")
    if not output:
        st.info(
            "No commands have been run in this session yet. Buttons on "
            "BOF Batch Builder land their output here."
        )
    else:
        label = st.session_state.get("last_command_label", "Last command")
        exit_code = st.session_state.get("last_command_exit_code", 0)
        elapsed = st.session_state.get("last_command_elapsed_s", 0.0)
        cmd = st.session_state.get("last_command_cmd", "")
        summary = st.session_state.get("last_command_summary", {}) or {}

        col_a, col_b, col_c = st.columns(3)
        col_a.metric("Command", label)
        col_b.metric("Exit code", str(exit_code))
        col_c.metric("Elapsed", f"{elapsed:.1f}s")

        if summary:
            st.write("**Parsed summary:**")
            st.dataframe(
                [{"metric": k, "value": v} for k, v in summary.items()],
                hide_index=True,
                use_container_width=True,
            )
        st.caption(f"`{cmd}`")
        st.code(output, language="text")

    st.divider()

    # --- Raw CLI buttons -----------------------------------------------
    st.subheader("Diagnostic commands")
    st.caption("These call `python main.py …` directly inside the container.")
    c1, c2, c3 = st.columns(3)
    if c1.button("Check Browser", key="adv_check_browser"):
        run_cli(["--check-browser"])
    if c2.button("Validate Manifest", key="adv_validate_manifest"):
        run_cli(["--validate-manifest", str(MANIFEST_PATH)])
    if c3.button("Load Manifest Fresh", key="adv_load_manifest_fresh",
                 disabled=not st.checkbox(
                     "Confirm fresh load (clears products.csv)",
                     value=False, key="adv_confirm_fresh",
                 )):
        run_cli(["--load-manifest", str(MANIFEST_PATH), "--fresh"])

    c4, c5, c6 = st.columns(3)
    if c4.button("Sync Favorites (legacy)", key="adv_sync_legacy",
                 help=(
                     "Updates CSV row statuses by matching Flow ❤️ to "
                     "captured media_ids. The main 'Scan Favorites' button "
                     "above uses the same command but presents results "
                     "differently."
                 )):
        run_cli(["--sync-favorites"])
    if c5.button("Generate Videos from Approved Rows (legacy)", key="adv_gen_videos_rows",
                 help=(
                     "Iterates CSV rows with status=image_approved + a "
                     "captured media_id. The main 'Generate Videos' button "
                     "iterates Flow's ❤️ tiles directly and doesn't need "
                     "this legacy path."
                 )):
        run_cli(["--generate-videos-from-approved-rows", "--limit", "30"])
    if c6.button("List Unmatched Favorites", key="adv_list_unmatched"):
        run_cli(["--list-unmatched-favorites"])

    st.divider()

    # --- Log files on disk ---------------------------------------------
    st.subheader("Log files on disk")
    if LOGS_DIR.exists():
        log_files = sorted(LOGS_DIR.glob("*.log"), reverse=True)
        if log_files:
            selected = st.selectbox(
                "Pick a log file", [f.name for f in log_files],
                key="adv_logfile_pick",
            )
            chosen = next((f for f in log_files if f.name == selected), None)
            if chosen:
                try:
                    text = chosen.read_text(encoding="utf-8")
                except OSError as exc:
                    st.error(f"Could not read {chosen}: {exc}")
                else:
                    st.caption(f"{chosen}  ·  {len(text):,} chars")
                    st.code(text[-50_000:], language="text")
        else:
            st.info("No log files yet.")
    else:
        st.info(f"Logs dir does not exist: {LOGS_DIR}")

    # Error screenshots (if any).
    if LOGS_DIR.exists():
        screenshots = sorted(LOGS_DIR.glob("*.png"), reverse=True)
        if screenshots:
            with st.expander(f"Error screenshots ({len(screenshots)})", expanded=False):
                for shot in screenshots[:12]:
                    st.caption(shot.name)
                    st.image(str(shot), use_container_width=True)

    st.divider()

    # --- Pointers to the old multi-page UI -----------------------------
    st.subheader("Legacy pages")
    st.caption(
        "The original page-by-page UI is still reachable if you need it. "
        "These do exactly what the BOF Batch Builder sections do — kept "
        "around for debugging unusual cases."
    )
    legacy_targets = {
        "Original Dashboard":         "Original Dashboard",
        "AI Product Intake (legacy)": "AI Product Intake (legacy)",
        "Reference Images (legacy)":  "Reference Images (legacy)",
        "Manifest Builder (legacy)":  "Manifest Builder (legacy)",
        "Batch Controls (legacy)":    "Batch Controls (legacy)",
        "Old per-step pages":         "1. Dashboard",
        "Old Settings page":          "Settings",
        "Old Logs page":              "7. Logs",
    }
    cols = st.columns(2)
    for i, (label, target) in enumerate(legacy_targets.items()):
        if cols[i % 2].button(label, key=f"adv_goto_{target}"):
            st.session_state["force_page"] = target
            st.rerun()


GUIDED_PAGES = {
    "BOF Batch Builder":    page_bof_batch_builder,
    "Setup":                page_setup,
    "Advanced / Logs":      page_advanced_logs,
}

# Old per-step pages kept reachable for debug, but no longer in the
# primary sidebar list. Reached via Advanced / Logs → Legacy pages.
ADVANCED_PAGES = {
    "1. Dashboard":               page_guided_dashboard,
    "2. Product Intake":          page_product_intake,
    "3. Images":                  page_images,
    "4. AI Prompts":              page_ai_prompts,
    "5. Export Manifest":         page_export_manifest,
    "6. Flow Batch Run":          page_flow_batch_run,
    "7. Logs":                    page_logs,
    "AI Product Intake (legacy)": page_ai_intake,
    "Reference Images (legacy)":  page_reference_images,
    "Manifest Builder (legacy)":  page_manifest,
    "Batch Controls (legacy)":    page_batch,
    "Original Dashboard":         page_dashboard,
    "Settings":                   page_settings,
}

# Combined dict for goto-page lookups (the dashboard's buttons reference
# guided-page labels only).
PAGES: dict[str, callable] = {**GUIDED_PAGES, **ADVANCED_PAGES}


def main() -> None:
    st.set_page_config(
        page_title="Flow BOF Automation",
        layout="wide",
    )

    # Honor any goto-page intent the previous run set, BEFORE the radio
    # widget is instantiated. Streamlit reads session_state[key] as the
    # widget's initial value when set this way.
    forced = st.session_state.pop("force_page", None)
    if forced and forced in PAGES:
        if forced in GUIDED_PAGES:
            st.session_state["sidebar_choice"] = forced
            # Also reset the advanced selector so it doesn't shadow the
            # guided choice on the next render.
            st.session_state["sidebar_advanced_choice"] = "(none)"
        elif forced in ADVANCED_PAGES:
            st.session_state["sidebar_advanced_choice"] = forced
    elif "sidebar_choice" not in st.session_state and not _ai_provider_is_configured():
        # First-launch experience: drop the user on Setup if no AI
        # provider is configured yet. Once they save settings, the next
        # session opens straight on BOF Batch Builder.
        st.session_state["sidebar_choice"] = "Setup"

    with st.sidebar:
        st.markdown("### Flow BOF Automation")
        choice = st.radio(
            "Workflow",
            list(GUIDED_PAGES.keys()),
            key="sidebar_choice",
            label_visibility="collapsed",
        )
        # Legacy per-step pages remain reachable but live inside a
        # collapsed expander labelled "Legacy / debug pages" so the
        # primary sidebar stays at three items as the spec requires.
        with st.expander("Legacy / debug pages", expanded=False):
            adv = st.radio(
                "Legacy",
                ["(none)", *ADVANCED_PAGES.keys()],
                key="sidebar_advanced_choice",
                label_visibility="collapsed",
            )
            if adv != "(none)":
                choice = adv
        st.divider()
        st.caption(
            "Keep this tab focused while a command runs. Raw logs are "
            "hidden behind 'Show technical logs' on each run."
        )

    PAGES[choice]()


if __name__ == "__main__":
    main()
